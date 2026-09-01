# ============================================================
# 1. IMPORTS & GLOBAL CONFIGURATION
# ============================================================
import os
import re
import io
import json
import math
import tempfile
import warnings
from datetime import datetime
from collections import Counter
from zoneinfo import ZoneInfo
from email.message import EmailMessage

import pandas as pd
import gradio as gr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

warnings.filterwarnings("ignore")

# Timezone Configuration (Pakistan Standard Time - PKT)
PKT = ZoneInfo("Asia/Karachi")

# Runtime Memory Incident Tracking Database
incident_records = {}

# RCBS Group Email Mapping for CC
RCBS_GROUPS = {
    "None": "",
    "RCBS North": "group.rcbs.north@gmail.com",
    "RCBS Central": "group.rcbs.central@gmail.com",
    "RCBS South": "group.rcbs.south@gmail.com",
}


# ============================================================
# 2. SHARED HELPERS & TIME FUNCTIONS
# ============================================================
def clean(value):
    """Clean and strip input string safely."""
    return (value or "").strip()


def get_current_time():
    """Returns current datetime in Asia/Karachi timezone."""
    return datetime.now(PKT)


def format_dt(value):
    """Formats datetime object to standard corporate representation."""
    if not value:
        return "NA"
    return value.strftime("%d-%m-%Y %I:%M %p")


def detect_service_type(text):
    """Detects standard telecom service type from label or complaint string."""
    value = clean(text).upper()

    checks = [
        ("BGP", "BGP DIA"),
        ("MPLS", "MPLS"),
        ("DPLC", "DPLC"),
        ("TURBONET", "Turbonet"),
        ("TURBO", "Turbonet"),
        ("SIP PRI", "SIP PRI"),
        ("SIP_PRI", "SIP PRI"),
        ("IPLC", "IPLC"),
        ("DARKCORE", "Darkcore Fiber"),
        ("DARK CORE", "Darkcore Fiber"),
        ("M2M", "M2M"),
        ("PRI", "PRI"),
        ("SIP", "SIP"),
        ("DIA", "DIA"),
    ]

    for token, service in checks:
        if token in value:
            return service

    return "Corporate"


def extract_service_id(label):
    """Extracts short unique service link identifier for concise subjects."""
    value = clean(label)
    if not value:
        return "Service"

    patterns = [
        r"\bDPLC\d+SL\d+\b",
        r"\bDIA\d+SL\d+\b",
        r"\bTurbo\d+SL\d+\b",
        r"\bLNK[A-Z0-9]+\b",
        r"\bMPLS[A-Z0-9-]*\b",
        r"\bPRI[A-Z0-9-]*\b",
        r"\bSIP[A-Z0-9-]*\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, value, flags=re.IGNORECASE)
        if match:
            return match.group(0)

    if "_" in value:
        candidate = value.split("_")[-1].strip()
        if candidate:
            return candidate[:70]

    return value[:70]


def priority_prefix(priority):
    """Returns email subject prefix tag according to escalation priority."""
    mapping = {
        "Normal": "",
        "Follow-up": "FOLLOW-UP | ",
        "Urgent": "URGENT | ",
        "Critical": "CRITICAL | ",
    }
    return mapping.get(priority, "")


def make_subject(label, topic, priority="Normal", ticket=""):
    """Standardized email subject generator."""
    service_id = extract_service_id(label)
    ticket = clean(ticket)
    ticket_part = f" | {ticket}" if ticket else ""
    return f"{priority_prefix(priority)}{topic} | {service_id}{ticket_part}"


def salutation(audience):
    """Returns appropriate professional salutation based on target audience."""
    if audience == "Customer":
        return "Dear Customer,"
    return "Dear Team,"


def calculate_duration(start_time, end_time):
    """Calculates formatted total outage duration between two datetime objects."""
    if not start_time or not end_time:
        return "NA"

    total_minutes = max(0, int((end_time - start_time).total_seconds() // 60))
    days, remainder = divmod(total_minutes, 1440)
    hours, minutes = divmod(remainder, 60)

    if days:
        return f"{days} Day(s) {hours} Hour(s) {minutes} Minute(s)"
    if hours:
        return f"{hours} Hour(s) {minutes} Minute(s)"
    return f"{minutes} Minute(s)"


def elapsed_text(start_time, end_time=None):
    """Calculates concise running age string for active incidents."""
    if not start_time:
        return "Not started"

    end_time = end_time or get_current_time()
    minutes = max(0, int((end_time - start_time).total_seconds() // 60))

    if minutes >= 1440:
        days, rem = divmod(minutes, 1440)
        hours, mins = divmod(rem, 60)
        return f"{days}d {hours}h {mins}m"
    if minutes >= 60:
        hours, mins = divmod(minutes, 60)
        return f"{hours}h {mins}m"
    return f"{minutes}m"


def incident_age_text(reported_time, added_time=None, status="OPEN"):
    """Generates visual status indicators based on active response age."""
    base_time = reported_time or added_time
    if not base_time:
        return "—"

    if status == "CLOSED":
        return "Closed"

    minutes = max(0, int((get_current_time() - base_time).total_seconds() // 60))

    if minutes < 5:
        marker = "🟢"
    elif minutes < 15:
        marker = "🟡"
    elif minutes < 30:
        marker = "🟠"
    else:
        marker = "🔴"

    prefix = "Queued " if not reported_time else ""
    return f"{marker} {prefix}{elapsed_text(base_time)}"


TIME_MODES = [
    "Automatic - Current Pakistan Time",
    "Manual Date/Time",
]


def parse_datetime_input(value):
    """Parses various date-time format strings into a PKT-aware datetime object."""
    value = clean(value)
    if not value:
        raise ValueError("Manual date/time is empty.")

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %I:%M:%S %p",
        "%Y-%m-%d %I:%M %p",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %I:%M:%S %p",
        "%d-%m-%Y %I:%M %p",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %I:%M:%S %p",
        "%d/%m/%Y %I:%M %p",
    ]

    for fmt in formats:
        try:
            parsed = datetime.strptime(value, fmt)
            return parsed.replace(tzinfo=PKT)
        except ValueError:
            continue

    raise ValueError(
        "Invalid date/time. Use YYYY-MM-DD HH:MM, DD-MM-YYYY HH:MM, or include AM/PM."
    )


def update_incident_stage(label, stage):
    """Updates the last stage marker of an existing active incident."""
    label = clean(label)
    if label in incident_records:
        incident_records[label]["last_stage"] = stage


def format_ettr(ettr_choice, custom_ettr):
    """Formats standardized Expected Time to Restore (ETTR) statements."""
    if ettr_choice == "Awaited":
        return "ETTR will be shared once received from the concerned team."
    if ettr_choice == "Not Available":
        return "Currently, no confirmed ETTR is available. Further updates will be shared accordingly."
    if ettr_choice == "Not Applicable":
        return ""
    if ettr_choice == "Custom":
        value = clean(custom_ettr)
        return (
            f"The tentative ETTR shared by the concerned team is {value}."
            if value
            else "ETTR will be shared once received from the concerned team."
        )
    return f"The tentative ETTR shared by the concerned team is {ettr_choice}."


def multiline_links(value):
    """Splits multiline input text into clean lists of individual link labels."""
    entries = []
    for raw in clean(value).splitlines():
        item = raw.strip(" \t-•")
        if item:
            entries.append(item)
    return entries


# ============================================================
# 3. COMPLAINT MANAGER ENGINE
# ============================================================
COMPLAINT_DEFAULT_ISSUES = [
    "Link is down.",
    "Service degradation.",
    "Intermittent connectivity.",
    "High latency observed.",
    "Packet loss observed.",
    "Slow browsing issue.",
    "Bandwidth issue.",
    "Website / application issue.",
    "Other / To be classified",
]


def get_complaint_labels(include_closed=True):
    """Retrieves list of runtime complaint labels sorted by addition time."""
    labels = []
    for label, record in incident_records.items():
        if not include_closed and record.get("status") == "CLOSED":
            continue
        labels.append(label)

    labels.sort(
        key=lambda x: incident_records[x].get("added_time") or get_current_time(),
        reverse=True,
    )
    return labels


def register_complaints(raw_labels, default_issue, ticket):
    """Registers multiple complaints into runtime memory queue simultaneously."""
    entries = multiline_links(raw_labels)
    if not entries:
        return (
            gr.Dropdown(choices=get_complaint_labels(), value=None),
            "⚠️ Paste one or more complaints/service labels, one per line.",
            refresh_dashboard("Queued + Open"),
        )

    now = get_current_time()
    added = []
    skipped = []
    restarted = []

    for label in entries:
        existing = incident_records.get(label)

        if existing and existing.get("status") in ("QUEUED", "OPEN"):
            skipped.append(label)
            continue

        if existing and existing.get("status") == "CLOSED":
            restarted.append(label)

        incident_records[label] = {
            "added_time": now,
            "reported_time": None,
            "restoration_time": None,
            "service_type": detect_service_type(label),
            "issue_type": clean(default_issue) or "Other / To be classified",
            "ticket": clean(ticket),
            "status": "QUEUED",
            "last_stage": "Complaint added - awaiting opening",
        }
        added.append(label)

    choices = get_complaint_labels()
    selected = added[0] if added else (choices[0] if choices else None)

    parts = []
    if added:
        parts.append(f"✅ Added {len(added)} complaint(s) to runtime queue.")
    if restarted:
        parts.append(
            f"♻️ {len(restarted)} previously closed label(s) started as new queued complaint(s)."
        )
    if skipped:
        parts.append(
            f"ℹ️ Skipped {len(skipped)} duplicate complaint(s) already QUEUED/OPEN."
        )

    return (
        gr.Dropdown(choices=choices, value=selected),
        "\n".join(parts),
        refresh_dashboard("Queued + Open"),
    )


def load_selected_complaint(selected_label):
    """Loads complaint details into active editing state."""
    label = clean(selected_label)
    if not label or label not in incident_records:
        return "", "", "⚠️ Please select a runtime complaint."

    record = incident_records[label]
    status = record.get("status", "")
    opened = (
        format_dt(record["reported_time"])
        if record.get("reported_time")
        else "Not opened yet"
    )
    closed = (
        format_dt(record["restoration_time"])
        if record.get("restoration_time")
        else "Not closed"
    )

    summary = f"""✅ Complaint loaded.
Status: {status}
Service Type: {record.get("service_type", "")}
Issue: {record.get("issue_type", "")}
Opening Time: {opened}
Closing Time: {closed}
Ticket: {record.get("ticket", "")}"""

    return label, record.get("ticket", ""), summary


def remove_selected_complaint(selected_label):
    """Deletes complaint record from active memory."""
    label = clean(selected_label)
    if not label or label not in incident_records:
        return (
            gr.Dropdown(choices=get_complaint_labels(), value=None),
            "⚠️ Please select a valid runtime complaint.",
            refresh_dashboard("Queued + Open"),
        )

    del incident_records[label]
    choices = get_complaint_labels()
    selected = choices[0] if choices else None

    return (
        gr.Dropdown(choices=choices, value=selected),
        "✅ Selected complaint removed from runtime memory.",
        refresh_dashboard("Queued + Open"),
    )


def refresh_complaint_selector():
    """Refreshes dropdown list choices of complaints."""
    choices = get_complaint_labels()
    return gr.Dropdown(
        choices=choices,
        value=choices[0] if choices else None,
    )


def get_complaint_labels_by_status(statuses=None):
    """Filters complaint labels by status (QUEUED, OPEN, CLOSED)."""
    labels = []
    allowed = set(statuses) if statuses else None

    for complaint_label, record in incident_records.items():
        status = record.get("status", "QUEUED")
        if allowed is not None and status not in allowed:
            continue
        labels.append(complaint_label)

    labels.sort(
        key=lambda x: incident_records[x].get("added_time") or get_current_time(),
        reverse=True,
    )
    return labels


def refresh_opening_complaint_selector():
    """Refreshes list choices for Opening Email tab."""
    choices = get_complaint_labels_by_status({"QUEUED", "OPEN"})
    return gr.Dropdown(
        choices=choices,
        value=choices[0] if choices else None,
    )


def refresh_closure_complaint_selector():
    """Refreshes list choices for Closure Email tab."""
    choices = get_complaint_labels_by_status({"OPEN"})
    return gr.Dropdown(
        choices=choices,
        value=choices[0] if choices else None,
    )


def resolve_complaint_label(selected_label, fallback_label):
    """Resolves label between selected complaint dropdown or textbox fallback."""
    return clean(selected_label) or clean(fallback_label)


# ============================================================
# 4. CLICK-ONLY DATE / CLOCK PICKER HELPERS
# ============================================================
AUTO_TIME_MODE = "Automatic - Current Pakistan Time"
MANUAL_TIME_MODE = "Manual - Mouse Clock Picker"

HOUR_CHOICES = [f"{i:02d}" for i in range(1, 13)]
MINUTE_CHOICES = [f"{i:02d}" for i in range(60)]
AMPM_CHOICES = ["AM", "PM"]


def _picker_date_to_date(value):
    """Helper to convert string/datetime date pickers into python date objects."""
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.astimezone(PKT).date() if value.tzinfo else value.date()

    if isinstance(value, (int, float)):
        return datetime.fromtimestamp(value, tz=PKT).date()

    raw = clean(value)
    if not raw:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%d-%m-%Y",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue

    raise ValueError("Please select a valid date from the calendar.")


def resolve_click_clock_time(mode, date_value, hour_value, minute_value, ampm_value):
    """Calculates precise datetime object from mouse clock controls."""
    if mode != MANUAL_TIME_MODE:
        return get_current_time()

    selected_date = _picker_date_to_date(date_value)
    if selected_date is None:
        raise ValueError("Please select the date from the calendar.")

    hour_text = clean(hour_value)
    minute_text = clean(minute_value)
    ampm = clean(ampm_value).upper()

    if not hour_text or not minute_text or ampm not in AMPM_CHOICES:
        raise ValueError("Please select hour, minute and AM/PM using the clock controls.")

    hour12 = int(hour_text)
    minute = int(minute_text)
    if not (1 <= hour12 <= 12):
        raise ValueError("Hour must be between 1 and 12.")
    if not (0 <= minute <= 59):
        raise ValueError("Minute must be between 00 and 59.")

    hour24 = hour12 % 12
    if ampm == "PM":
        hour24 += 12

    return datetime(
        selected_date.year,
        selected_date.month,
        selected_date.day,
        hour24,
        minute,
        0,
        tzinfo=PKT,
    )


def clock_card_html(hour_value=None, minute_value=None, ampm_value=None, date_value=None):
    """Renders dynamic analog clock SVG preview card with digital time display."""
    now = get_current_time()

    try:
        hour12 = int(clean(hour_value)) if clean(hour_value) else ((now.hour - 1) % 12) + 1
    except ValueError:
        hour12 = ((now.hour - 1) % 12) + 1

    try:
        minute = int(clean(minute_value)) if clean(minute_value) else now.minute
    except ValueError:
        minute = now.minute

    ampm = clean(ampm_value).upper() if clean(ampm_value) else now.strftime("%p")
    if ampm not in AMPM_CHOICES:
        ampm = now.strftime("%p")

    minute_angle = minute * 6
    hour_angle = (hour12 % 12) * 30 + minute * 0.5

    def endpoint(cx, cy, length, angle_deg):
        angle = math.radians(angle_deg - 90)
        return cx + length * math.cos(angle), cy + length * math.sin(angle)

    cx, cy = 92, 72
    hx, hy = endpoint(cx, cy, 30, hour_angle)
    mx, my = endpoint(cx, cy, 43, minute_angle)

    nums = []
    for n in range(1, 13):
        x, y = endpoint(cx, cy, 55, n * 30)
        nums.append(
            f'<text x="{x:.1f}" y="{y + 4:.1f}" text-anchor="middle" '
            'font-size="11" font-weight="700" fill="#fff">'
            f'{n}</text>'
        )

    date_text = ""
    try:
        d = _picker_date_to_date(date_value)
        if d:
            date_text = d.strftime("%d-%m-%Y")
    except Exception:
        date_text = ""

    digital = f"{hour12:02d}:{minute:02d} {ampm}"
    subtitle = f"{date_text} · Pakistan (PKT)" if date_text else "Pakistan (PKT)"

    return f'''<div style="display:flex;align-items:center;gap:24px;background:#191919;color:white;border-radius:22px;padding:16px 24px;max-width:520px;min-height:150px;box-sizing:border-box;">
      <div style="min-width:190px;">
        <div style="font-size:30px;font-weight:750;line-height:1.1;">{digital}</div>
        <div style="font-size:14px;color:#c9c9c9;margin-top:10px;">{subtitle}</div>
        <div style="font-size:12px;color:#9fdaff;margin-top:8px;">Mouse clock preview</div>
      </div>
      <svg width="184" height="144" viewBox="0 0 184 144" role="img" aria-label="Selected analogue time">
        <circle cx="92" cy="72" r="67" fill="#202020" stroke="#3a3a3a" stroke-width="2"/>
        {''.join(nums)}
        <line x1="92" y1="72" x2="{hx:.1f}" y2="{hy:.1f}" stroke="#e8e8e8" stroke-width="4" stroke-linecap="round"/>
        <line x1="92" y1="72" x2="{mx:.1f}" y2="{my:.1f}" stroke="#e8e8e8" stroke-width="3" stroke-linecap="round"/>
        <circle cx="92" cy="72" r="4" fill="#ff6b35"/>
      </svg>
    </div>'''


def current_clock_picker_values():
    """Resets clock picker state variables to current PKT time."""
    now = get_current_time().replace(second=0, microsecond=0)
    return (
        MANUAL_TIME_MODE,
        now.strftime("%Y-%m-%d"),
        now.strftime("%I"),
        now.strftime("%M"),
        now.strftime("%p"),
        clock_card_html(
            now.strftime("%I"),
            now.strftime("%M"),
            now.strftime("%p"),
            now.strftime("%Y-%m-%d"),
        ),
    )


def update_clock_preview(date_value, hour_value, minute_value, ampm_value):
    """Wrapper to update clock preview on dropdown value changes."""
    return clock_card_html(hour_value, minute_value, ampm_value, date_value)


# ============================================================
# 5. OPENING EMAIL ENGINE
# ============================================================
OPENING_ISSUES = [
    "Link is down.",
    "Service degradation.",
    "Intermittent connectivity.",
    "High latency observed.",
    "Packet loss observed.",
    "Slow browsing issue.",
    "Bandwidth issue.",
    "Website / application issue.",
    "Other / Manual",
]


def generate_opening_email(
    selected_complaint,
    label,
    issue_type,
    custom_issue,
    ticket,
    priority,
    opening_time_mode,
    opening_date,
    opening_hour,
    opening_minute,
    opening_ampm,
):
    """Generates standardized Opening Acknowledgement email template."""
    label = resolve_complaint_label(selected_complaint, label)
    if not label:
        return "", "", "", "⚠️ Please select a complaint or enter the service label."

    issue = (
        clean(custom_issue) or "Service issue reported."
        if issue_type == "Other / Manual"
        else issue_type
    )

    try:
        requested_time = resolve_click_clock_time(
            opening_time_mode,
            opening_date,
            opening_hour,
            opening_minute,
            opening_ampm,
        )
    except (ValueError, TypeError) as exc:
        return label, "", "", f"⚠️ Invalid opening date/time: {exc}"

    service_type = detect_service_type(label)
    now = get_current_time()
    record = incident_records.get(label)
    reused_existing_time = False

    if record and record.get("status") == "OPEN" and record.get("reported_time"):
        reported_time = record["reported_time"]
        reused_existing_time = True
    else:
        reported_time = requested_time
        incident_records[label] = {
            "added_time": record.get("added_time") if record else now,
            "reported_time": reported_time,
            "restoration_time": None,
            "service_type": (
                record.get("service_type")
                if record and record.get("service_type")
                else service_type
            ),
            "issue_type": issue,
            "ticket": clean(ticket) or (record.get("ticket", "") if record else ""),
            "status": "OPEN",
            "last_stage": "Initial response",
        }

    record = incident_records[label]
    record["issue_type"] = issue
    record["ticket"] = clean(ticket) or record.get("ticket", "")
    record["last_stage"] = "Initial response"
    record["status"] = "OPEN"
    record["restoration_time"] = None

    subject = make_subject(label, issue.rstrip("."), priority, record.get("ticket", ""))

    email = f"""Dear Customer,

We acknowledge receipt of your complaint regarding service degradation/impact on your {record.get("service_type", service_type)} service. Our Corporate NOC has initiated an investigation.

Details:
Service Details: {label}
Issue Type: {issue}
Incident Start Time: {format_dt(reported_time)}
Current Status: Under investigation
Reference Ticket: {record.get("ticket", "")}

Our technical teams are actively working to identify the root cause and restore service at the earliest possible time. Regular updates will be shared until the issue is fully resolved.

We appreciate your patience and cooperation."""

    if reused_existing_time:
        runtime = (
            f"ℹ️ Complaint was already OPEN. Existing opening time retained: "
            f"{format_dt(reported_time)}."
        )
    else:
        source_note = (
            "manual mouse clock"
            if opening_time_mode == MANUAL_TIME_MODE
            else "current Pakistan time"
        )
        runtime = (
            f"✅ Complaint OPENED at {format_dt(reported_time)} "
            f"for {extract_service_id(label)} ({source_note})."
        )

    return label, subject, email, runtime


# ============================================================
# 6. CUSTOMER-END & FINDINGS ENGINE
# ============================================================
FINDING_OPTIONS = [
    "No problematic alarm observed at Node end",
    "Port is up",
    "Tunnel is intact",
    "ARP is not observed",
    "MAC is not observed",
    "User is offline at BRAS",
    "Optical power is optimal",
    "Traffic is observed at our end",
    "Latency is optimal",
    "No abnormality observed in transmission path",
]

CUSTOMER_ACTIONS = [
    "Verify last-mile media",
    "Check CPE/router and power",
    "Check SFP/optical module and equipment port",
    "Reconnect the affected PPPoE user",
    "Test through standalone laptop/device",
    "Bypass LAN and perform direct testing",
    "Share active POC for joint troubleshooting",
    "Share problematic stats for further investigation",
]


def finding_bullet(finding, service_type):
    if finding == "Traffic is observed at our end" and service_type == "Turbonet":
        return "Traffic is being observed at our end."
    return finding + "."


def customer_action_sentence(action):
    mapping = {
        "Verify last-mile media": "Kindly verify the service from the last-mile media and share your feedback.",
        "Check CPE/router and power": "Kindly verify the CPE/router status and power availability at your end.",
        "Check SFP/optical module and equipment port": "Kindly check the SFP/optical module and equipment port at your end and share updated status.",
        "Reconnect the affected PPPoE user": "Kindly reconnect the affected PPPoE user and confirm whether the session comes online.",
        "Test through standalone laptop/device": "Kindly perform testing through a standalone laptop/device to isolate any internal LAN impact.",
        "Bypass LAN and perform direct testing": "Kindly bypass the internal LAN and perform direct testing from the edge device.",
        "Share active POC for joint troubleshooting": "Kindly share an active POC so joint troubleshooting can be arranged.",
        "Share problematic stats for further investigation": "If the issue persists, kindly share the problematic statistics for further investigation.",
    }
    return mapping.get(action, "")


def generate_customer_end_email(
    label,
    issue_summary,
    findings,
    requested_action,
    priority,
    ticket,
):
    """Generates response email when issue is identified at customer side."""
    label = clean(label)
    if not label:
        return "", "⚠️ Please enter the service label.", ""

    service_type = detect_service_type(label)
    subject = make_subject(
        label,
        issue_summary or "Customer-End Verification",
        priority,
        ticket,
    )

    lines = [
        "Dear Customer,",
        "",
        "We have thoroughly checked the reported service.",
        "",
        "Findings are shared below:",
    ]

    if findings:
        for item in findings:
            lines.append(f"• {finding_bullet(item, service_type)}")
    else:
        lines.append("• No abnormality is currently observed at our end.")

    action_line = customer_action_sentence(requested_action)
    if action_line:
        lines.extend(["", action_line])

    lines.extend([
        "",
        "Please share your feedback so we can proceed accordingly.",
        "",
        "Your cooperation in this matter will be highly appreciated.",
    ])

    policy = ""
    if service_type == "Turbonet":
        policy = (
            "⚠️ Internal Turbo policy: Do not share the Turbonet graph with "
            "the customer. For graph-related requests, route the matter through "
            "KAM/GCSSQ."
        )

    return subject, "\n".join(lines), policy


# ============================================================
# 7. VENDOR & INTERNAL ESCALATION ENGINE (CC ROUTING & MANUAL OVERRIDE)
# ============================================================
ESCALATION_TARGETS = [
    "Netsat",
    "Comstar",
    "VT",
    "GCS",
    "WISS",
    "Nexlinx",
    "Connect",
    "Other",
]

ESCALATION_LEVELS = [
    "Initial engagement",
    "Follow-up",
    "Urgent",
    "Critical",
    "Multiple follow-ups / no response",
    "Request exact delay reason / RFO",
]

ESCALATION_FORMAT_OPTIONS = [
    "Option A (Standard Matrix Format)",
    "Option B (Subject-Linked Format)",
    "Option C ([Ticket] : [Subject] : [Vendor])",
]


def get_vendor_l1_l2_emails(vendor_name):
    """Extracts only Level 1 and Level 2 emails for a vendor (semicolon separated)."""
    if not vendor_name or vendor_name not in vendors_matrix_db:
        return ""
    emails = []
    for row in vendors_matrix_db[vendor_name]:
        lvl = str(row.get("Level", "")).upper()
        if "LEVEL 1" in lvl or "LEVEL 2" in lvl or "L1" in lvl or "L2" in lvl:
            raw_email = row.get("Email", "")
            for e in re.split(r"[;, ]+", raw_email):
                clean_e = e.strip()
                if clean_e and clean_e not in emails:
                    emails.append(clean_e)
    return "; ".join(emails)


def generate_outlook_eml_file(to_emails, cc_emails, subject, body_text, matrix_html):
    """Generates a standard .eml file with CC header that opens directly in Microsoft Outlook."""
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = "noc@company.com"
    if to_emails:
        msg["To"] = to_emails
    if cc_emails:
        msg["Cc"] = cc_emails

    msg.set_content(body_text)



def generate_escalation_email(
    label,
    target,
    escalation_level,
    escalation_format,
    rcbs_group,
    custom_to,
    custom_cc,
    custom_c_subject,
    ticket,
    findings,
    requested_action,
    priority,
    custom_target="",
):
    """Generates vendor / internal team escalation email across Option A, B, and C formats,

    and auto-creates a downloadable Outlook (.eml) message file with proper CC headers.
    """
    label_clean = clean(label)
    target_name = custom_target.strip() if target == "Other" and custom_target else target
    if not target_name:
        target_name = "Vendor"

    # 1. Resolve To & CC (User input takes priority; fallback to auto-fetch)
    to_emails = clean(custom_to) or get_vendor_l1_l2_emails(target_name)
    cc_emails = clean(custom_cc) or RCBS_GROUPS.get(rcbs_group, "")

    # 2. Determine Subject based on Option A, B, C
    if "Option A" in escalation_format:
        topic = f"{escalation_level} - {target_name}"
        subject = make_subject(label_clean, topic, priority, ticket)
        intro_sentence = "Please check the below-mentioned service, as the customer is currently facing a connectivity issue."
    elif "Option B" in escalation_format:
        subject = f"{label_clean} || {target_name}" if label_clean else f"Service Issue || {target_name}"
        intro_sentence = "Please check the link mentioned in subject, as the customer is currently facing a connectivity issue."
    else:
        # Option C: [Ticket Number] : [Subject] : [Vendor Name]
        sub_text = clean(custom_c_subject) or label_clean or "Service Outage"
        tkt_text = clean(ticket) or "No-Ticket"
        subject = f"[{tkt_text}] : [{sub_text}] : [{target_name}]"
        intro_sentence = "Please check the below-mentioned service, as the customer is currently facing a connectivity issue."

    # 3. Assemble Body
    greeting = f"Dear {target_name} Team,"
    extra_parts = []

    intro_map = {
        "Urgent": "Kindly prioritize the reported issue on urgent basis. The corporate customer is currently experiencing service impact/outage.",
        "Critical": "The reported issue is critically impacting our corporate customer. Immediate engagement and restoration on top priority are required.",
        "Multiple follow-ups / no response": f"Despite repeated follow-ups, we have not received a progressive update from the {target_name} team. The continued delay is impacting customer communication and is not acceptable.",
        "Request exact delay reason / RFO": "We have repeatedly requested a specific reason for the delay / clear RFO; however, only generic updates have been received. Kindly share the exact reason, defined action plan, and expected restoration timeline.",
    }
    if escalation_level in intro_map:
        extra_parts.append(intro_map[escalation_level])

    if findings:
        extra_parts.append("\nCurrent Findings:")
        if isinstance(findings, str):
            for line in findings.splitlines():
                if line.strip():
                    extra_parts.append(f"• {line.strip()}")
        elif isinstance(findings, (list, tuple)):
            for item in findings:
                if item:
                    extra_parts.append(f"• {item}.")

    req_act_clean = clean(requested_action)
    if req_act_clean:
        extra_parts.append(f"\nRequired Action: {req_act_clean}")

    if escalation_level in {
        "Urgent",
        "Critical",
        "Multiple follow-ups / no response",
        "Request exact delay reason / RFO",
    }:
        extra_parts.append("\nKindly share a clear progressive update/action plan at the earliest.")

    extra_content_str = "\n".join(extra_parts)
    extra_section = f"\n{extra_content_str}\n" if extra_parts else "\n"

    body = f"""{greeting}

{intro_sentence}

Service Details: {label_clean or "NA"}

Kindly investigate the issue, restore the service on priority, and share the ETTR at the earliest.{extra_section}
Your prompt support and cooperation will be highly appreciated."""

    # 4. Render HTML Matrix Table
    matrix_html = render_escalation_matrix_html(target_name)

    # 5. Generate Outlook .eml File
    eml_file_path = generate_outlook_eml_file(to_emails, cc_emails, subject, body, matrix_html)

    return subject, to_emails, cc_emails, body, eml_file_path, matrix_html


# ============================================================
# 8. CLOSURE & RFO ENGINE
# ============================================================
RCA_OPTIONS = [
    "Fiber Break (Dual/Triple)",
    "Fiber Break (Single)",
    "Fiber Break (Spur)",
    "Power Issue",
    "Power Issue at PTN / Node",
    "Port Down",
    "Port Problematic",
    "Customer related issue",
    "Customer own last mile",
    "Customer end power issue",
    "Hardware Faulty",
    "Configurations Issue",
    "Link Choking / Over-utilization",
    "Frequency Interference",
    "Optical Power Issue",
    "Vendor / Upstream Issue",
    "Routing Issue",
    "Submarine Cable Cut",
    "No Issue Observed",
    "Other / Manual",
]

ISSUE_FOUND_AT = [
    "Customer",
    "CONVEX",
    "NSS TEAM",
    "PRODUCT TEAM",
    "SELLER",
    "NOMC - Transmission Optical",
    "Vendor",
    "Region - Field Operations (FOPs)",
    "NOMC - IP Core",
    "Uplink ISP",
    "NOMC - Transmission Microwave",
    "TXN PM",
    "TXN CM",
    "RCA Awaited",
    "Other",
]

AUTO_ACTIONS = {
    "Fiber Break (Dual/Triple)": "Affected fiber section was restored/spliced and service connectivity was normalized.",
    "Fiber Break (Single)": "Affected fiber section was restored/spliced and service connectivity was normalized.",
    "Fiber Break (Spur)": "Spur fiber fault was restored and service connectivity was normalized.",
    "Power Issue": "Power was restored at the affected site/node and service became operational.",
    "Power Issue at PTN / Node": "Power was restored at the affected PTN/node and service became operational.",
    "Port Down": "Port connectivity was restored and the service became operational.",
    "Port Problematic": "The problematic port/connectivity was rectified and service was restored.",
    "Customer related issue": "Service became operational after customer-end verification/restoration.",
    "Customer own last mile": "The customer/last-mile issue was rectified and service became operational.",
    "Customer end power issue": "Customer-end power was restored and service became operational.",
    "Hardware Faulty": "The faulty hardware/equipment was rectified/replaced and service was restored.",
    "Configurations Issue": "Required configuration was rectified and service was restored.",
    "Link Choking / Over-utilization": "Bandwidth/utilization was normalized and the service was verified.",
    "Frequency Interference": "Wireless parameters/path were optimized and service stability was restored.",
    "Optical Power Issue": "Optical parameters were normalized and service connectivity was restored.",
    "Vendor / Upstream Issue": "The upstream/vendor issue was resolved and service was restored.",
    "Routing Issue": "Routing/path configuration was rectified and service was restored.",
    "Submarine Cable Cut": "Upstream capacity/path was restored or stabilized after the submarine cable incident.",
    "No Issue Observed": "No activity was performed at our end; the service was already found operational.",
}


def generate_closure_email(
    selected_complaint,
    label,
    issue_found_at,
    custom_issue_found_at,
    root_cause_option,
    custom_root_cause,
    corrective_action_option,
    custom_corrective_action,
    final_status,
    priority,
    closure_time_mode,
    closure_date,
    closure_hour,
    closure_minute,
    closure_ampm,
):
    """Generates closure RFO email and calculates total impact duration automatically."""
    label = resolve_complaint_label(selected_complaint, label)
    if not label:
        return "", "", "", "⚠️ Please select an OPEN complaint or enter the service label."

    record = incident_records.get(label)
    service_type = (
        record.get("service_type")
        if record and record.get("service_type")
        else detect_service_type(label)
    )

    root_cause = (
        clean(custom_root_cause) or "No issue observed"
        if root_cause_option == "Other / Manual"
        else root_cause_option
    )

    found_at = (
        clean(custom_issue_found_at)
        if issue_found_at == "Other"
        else issue_found_at
    )

    if corrective_action_option == "Auto from Root Cause":
        corrective_action = AUTO_ACTIONS.get(
            root_cause,
            "Required corrective action was completed and service was restored.",
        )
    elif corrective_action_option == "Other / Manual":
        corrective_action = (
            clean(custom_corrective_action)
            or "Required corrective action was completed."
        )
    else:
        corrective_action = corrective_action_option

    try:
        restoration_time = resolve_click_clock_time(
            closure_time_mode,
            closure_date,
            closure_hour,
            closure_minute,
            closure_ampm,
        )
    except (ValueError, TypeError) as exc:
        return label, "", "", f"⚠️ Invalid closing date/time: {exc}"

    no_issue_root_causes = [
        "No Issue Observed",
        "No Issue Found after testing",
    ]

    customer_end_indicators = [
        "Customer related issue",
        "Customer own last mile",
        "Customer end power issue",
        "Customer",
    ]

    is_no_issue_at_noc = False

    if root_cause in no_issue_root_causes:
        is_no_issue_at_noc = True

    if root_cause_option == "Other / Manual" and clean(custom_root_cause):
        custom_rc_lower = clean(custom_root_cause).lower()
        if any(phrase in custom_rc_lower for phrase in ["no issue", "no fault", "no abnormality", "no problem"]):
            is_no_issue_at_noc = True

    if found_at in customer_end_indicators:
        is_no_issue_at_noc = True

    no_action_phrases = [
        "No activity performed at our end",
        "no activity was performed",
        "already found operational",
        "already operating normally",
    ]
    if any(phrase in corrective_action.lower() for phrase in no_action_phrases):
        is_no_issue_at_noc = True

    if record and record.get("reported_time"):
        reported_time = record["reported_time"]

        if restoration_time < reported_time:
            return (
                label,
                "",
                "",
                f"⚠️ Closing time ({format_dt(restoration_time)}) cannot be earlier than "
                f"the saved opening time ({format_dt(reported_time)}).",
            )

        formatted_reported = format_dt(reported_time)

        if is_no_issue_at_noc:
            duration = "NA"
        else:
            duration = calculate_duration(reported_time, restoration_time)

        record["status"] = "CLOSED"
        record["last_stage"] = "Closed"
        record["restoration_time"] = restoration_time
        record["root_cause"] = root_cause
        record["issue_found_at"] = found_at
        record["corrective_action"] = corrective_action
    else:
        formatted_reported = "NA"
        duration = "NA"

        if record:
            record["status"] = "CLOSED"
            record["last_stage"] = "Closed without saved opening time"
            record["restoration_time"] = restoration_time

    subject = make_subject(label, "Service Restored / Closure", priority)

    if is_no_issue_at_noc:
        email = f"""Dear Customer,

We have thoroughly checked the reported service and no abnormality is currently observed at our end.

Incident Summary:
Service Details: {label}
Issue Found At: {found_at or "NA"}
Root Cause: {root_cause}
Corrective Action: {corrective_action}
Reported Time: {formatted_reported}
Verification Time: {format_dt(restoration_time)}
Service Impact Duration: NA

Note: As no issue was identified within our network, the service impact duration is not applicable.

The service is currently operating normally. Kindly verify and confirm if performance is satisfactory at your end.

We regret any inconvenience caused and appreciate your cooperation.

{final_status}"""
    else:
        email = f"""Dear Customer,

We are pleased to inform you that your {service_type} service is fully operational.

Incident Summary:
Service Details: {label}
Issue Found At: {found_at or "NA"}
Root Cause: {root_cause}
Corrective Action: {corrective_action}
Reported Time: {formatted_reported}
Restoration Time: {format_dt(restoration_time)}
Service Impact Duration: {duration}

The service is now operating normally. Kindly verify and confirm if performance is satisfactory at your end.

We regret any inconvenience caused and appreciate your cooperation throughout the restoration process.

{final_status}"""

    if record and record.get("reported_time"):
        if is_no_issue_at_noc:
            runtime = (
                f"✅ Complaint CLOSED at {format_dt(restoration_time)}. "
                f"Service Impact Duration: NA (No issue found at NOC end)."
            )
        else:
            runtime = (
                f"✅ Complaint CLOSED at {format_dt(restoration_time)}. "
                f"Total impact duration: {duration}."
            )
    elif record:
        runtime = (
            f"ℹ️ Complaint CLOSED at {format_dt(restoration_time)}, "
            "but no opening time was saved, so duration is NA."
        )
    else:
        runtime = (
            "ℹ️ Closure generated without a runtime complaint record. "
            "Reported Time and Duration are NA."
        )

    return label, subject, email, runtime


# ============================================================
# 9. STATS & TROUBLESHOOTING ENGINE
# ============================================================
STATS_SCENARIOS = [
    "Packet Loss / High Latency",
    "Slow Speed / Throughput",
    "Turbo Slow Speed",
    "Website Issue",
    "Particular Website / URL Issue",
    "Banking Application Issue",
    "TikTok Issue",
    "WhatsApp / Meta Issue",
    "PPPoE / BRAS Issue",
    "Voice / PRI / SIP Issue",
    "Optical Power Issue",
    "Routing / BGP Issue",
]


def stats_items(scenario):
    mapping = {
        "Packet Loss / High Latency": [
            "Ping statistics from the affected user",
            "WinMTR / traceroute towards the problematic destination",
            "Source IP / testing IP pool",
            "Destination IP / URL",
            "Affected user details",
            "Exact issue occurrence date and time",
        ],
        "Slow Speed / Throughput": [
            "Speedtest result",
            "Speedtest server details",
            "Affected user/source IP",
            "Direct/standalone testing result after bypassing the LAN",
            "Current issue occurrence time",
        ],
        "Turbo Slow Speed": [
            "Direct speed and browsing test through a standalone laptop/device",
            "Affected PPPoE user / source IP",
            "Speedtest server details and result",
            "Issue occurrence time",
        ],
        "Website Issue": [
            "Source IP pool",
            "Browser error screenshot",
            "Traceroute / WinMTR towards the problematic destination",
            "NSLOOKUP result of the problematic destination",
            "NSLOOKUP result of whoami.akamai.com",
            "Tracetcp towards the affected URL on port 443",
            "Primary and secondary DNS configuration",
        ],
        "Particular Website / URL Issue": [
            "Snapshot of whatismyipaddress.com",
            "Source IP pool",
            "Browser error screenshot",
            "Traceroute towards the problematic destination",
            "WinMTR / tracetcp towards the affected URL on port 443",
            "NSLOOKUP result of the problematic destination",
            "Primary and secondary DNS configuration",
        ],
        "Banking Application Issue": [
            "Snapshot of whatismyipaddress.com",
            "Source IP pool",
            "Browser/application error screenshot",
            "NSLOOKUP / traceroute / WinMTR details where applicable",
            "PCAP trace from the affected Android phone using PCAPdroid",
        ],
        "TikTok Issue": [
            "DNS server IP configured at the edge device",
            "NSLOOKUP whoami.akamai.com screenshot",
            "Standalone test using Ethernet cable and a test PPPoE user",
            "Windows Task Manager > Performance > Ethernet screenshot while playing a TikTok video",
            "Affected user/source IP and issue occurrence time",
        ],
        "WhatsApp / Meta Issue": [
            "Affected user/source IP",
            "Issue occurrence time",
            "Ping / WinMTR towards a reachable problematic destination if available",
            "DNS configuration",
            "Standalone testing result after bypassing the LAN",
        ],
        "PPPoE / BRAS Issue": [
            "PPPoE username",
            "Affected user MAC address",
            "CPE/router status",
            "Exact disconnect/login time",
            "Standalone dial/test result",
            "Screenshot/error message if authentication fails",
        ],
        "Voice / PRI / SIP Issue": [
            "Party-A number",
            "Party-B number",
            "Exact call date and time",
            "Incoming / outgoing call scenario",
            "Observed error / announcement",
            "CLI displayed at Party-B where applicable",
        ],
        "Optical Power Issue": [
            "Current TX/RX optical power readings",
            "SFP/optical module details",
            "Equipment port status",
            "Patch-cord/fiber inspection result",
            "Relevant alarm/screenshot for reference",
        ],
        "Routing / BGP Issue": [
            "Source IP pool / affected prefix",
            "Problematic destination IP / URL",
            "Traceroute / WinMTR",
            "Current route advertisement details",
            "Expected upstream/path if known",
        ],
    }
    return mapping.get(scenario, ["Relevant problematic statistics and timestamps"])


def generate_stats_request(label, scenario, audience, custom_context):
    """Generates required technical stats email request."""
    label = clean(label)
    subject = make_subject(
        label or scenario,
        f"Required Stats - {scenario}",
        "Normal",
    )

    lines = [
        salutation(audience),
        "",
        "To proceed with detailed investigation of the reported issue, kindly share the following information/statistics:",
        "",
    ]

    for item in stats_items(scenario):
        lines.append(f"• {item}")

    if clean(custom_context):
        lines.extend(["", clean(custom_context)])

    if scenario == "Banking Application Issue":
        lines.extend([
            "",
            "PCAPdroid capture method:",
            "Select PCAP file → select the target application → START capture → reproduce the banking-app issue/login error → STOP capture → share the generated PCAP file.",
        ])

    lines.extend([
        "",
        "Once the required information is received, we will proceed with further investigation accordingly.",
        "",
        "Your cooperation is highly appreciated.",
    ])

    policy = ""
    if scenario == "Turbo Slow Speed":
        policy = (
            "⚠️ Internal Turbo policy: Never share the Turbonet graph directly "
            "with the customer. If the customer requests the graph, route the "
            "request to KAM; graph-policy escalation should be handled with GCSSQ."
        )

    return subject, "\n".join(lines), policy


# ============================================================
# 10. PROGRESS UPDATE ENGINE
# ============================================================
PROGRESS_OPTIONS = [
    "Concerned team engaged",
    "Fault localization in progress",
    "Team dispatched",
    "Team reached site",
    "Fiber splicing in progress",
    "Rerouting in progress",
    "Configuration activity in progress",
    "Vendor engaged",
    "Customer coordination required",
    "Testing in progress",
    "Monitoring link stability",
]

PROGRESS_TEXT = {
    "Concerned team engaged": "Our concerned team is actively engaged and working on the reported issue.",
    "Fault localization in progress": "Fault localization is currently in progress and is being followed up on priority.",
    "Team dispatched": "The concerned field team has been dispatched for onsite troubleshooting/restoration.",
    "Team reached site": "The field team has reached the site and troubleshooting/restoration activity is in progress.",
    "Fiber splicing in progress": "Fiber splicing activity is currently in progress at the affected section.",
    "Rerouting in progress": "Rerouting activity is in progress to restore the affected service through an alternate path.",
    "Configuration activity in progress": "The concerned technical team is carrying out the required configuration activity.",
    "Vendor engaged": "The concerned vendor/upstream team has been engaged and is investigating the issue.",
    "Customer coordination required": "Further troubleshooting requires customer-side coordination. Kindly ensure an active POC is available.",
    "Testing in progress": "Detailed testing is currently in progress with the concerned teams.",
    "Monitoring link stability": "The link is currently under monitoring to verify stability and service performance.",
}


def generate_progress_email(
    label,
    progress_status,
    ettr_choice,
    custom_ettr,
    custom_note,
    priority,
    audience,
):
    """Generates progress status update email."""
    label = clean(label)
    if not label:
        return "", "⚠️ Please enter the service label."

    update_incident_stage(label, progress_status)
    subject = make_subject(label, f"Progress Update - {progress_status}", priority)

    body = f"""{salutation(audience)}

Please find the latest update regarding the reported service:

{PROGRESS_TEXT.get(progress_status, progress_status)}"""

    if clean(custom_note):
        body += f"\n\n{clean(custom_note)}"

    ettr_line = format_ettr(ettr_choice, custom_ettr)
    if ettr_line:
        body += f"\n\n{ettr_line}"

    body += (
        "\n\nFurther updates will be shared accordingly."
        "\n\nYour patience and cooperation are highly appreciated."
    )

    return subject, body


# ============================================================
# 11. OUTAGE ANALYZER ENGINE (STREAMLINED BGP MATCHING)
# ============================================================
outage_history = []


def extract_bw_numeric(label):
    match = re.search(r'(\d+(?:\.\d+)?)\s*(?:M|G|K)BPS', str(label).replace('Mpbs', 'Mbps'), re.IGNORECASE)
    if match:
        val = float(match.group(1))
        if 'G' in match.group(0).upper():
            val *= 1000
        return val
    return 0


def extract_bw_text(label):
    match = re.search(r'(\d+(?:\.\d+)?)\s*((?:M|G|K)BPS)', str(label).replace('Mpbs', 'Mbps'), re.IGNORECASE)
    if match:
        return f"{match.group(1)} {match.group(2).upper()}"
    return "NA"


def detect_category(label):
    """Streamlined category detection separating BGP DIA from DIA and transport."""
    lbl = str(label).upper()

    # 1. BGP DIA (Prioritized: matches any link containing BGP)
    if "BGP" in lbl:
        return "BGPDIA"

    # 2. Broadband / Access
    if "TURBO" in lbl or "TURBONET" in lbl:
        return "Turbonet"

    # 3. Transport & Dedicated Circuits
    if "MPLS" in lbl or "VPLS" in lbl:
        return "MPLS"
    if "DPLC" in lbl or "EPL" in lbl:
        return "DPLC"
    if "IPLC" in lbl:
        return "IPLC"
    if "M2M" in lbl:
        return "M2M"

    # 4. Standard DIA
    if "DIA" in lbl:
        return "DIA"

    # 5. Voice / Signaling
    if "SIP" in lbl or "PRI" in lbl:
        return "SIP/PRI"

    return "Other"


def extract_client_name(label):
    parts = [p.strip() for p in str(label).strip().split("_") if p.strip()]
    if parts and "ESSCLIENT" in parts[0].upper():
        parts = parts[1:]
    bw_idx = next((i for i, p in enumerate(parts) if re.search(r'\d+(?:\.\d+)?\s*(?:M|G|K)BPS', p.replace('Mpbs', 'Mbps'), re.I)), None)
    c_parts = parts[:bw_idx] if bw_idx is not None else (parts[:-1] if len(parts) > 1 else parts)
    control_words = {"DIA", "MPLS", "DPLC", "M2M", "TURBO", "TURBONET", "BGP", "BGPDIA", "DIABGP", "SIP", "PRI", "SIPPRI", "CENTRAL", "NORTH", "SOUTH"}
    clean_parts = [p for p in c_parts if p.upper() not in control_words and not p.isdigit()]
    return " ".join(clean_parts[:-1] if len(clean_parts) >= 2 else clean_parts).strip()


def extract_unique_id(label):
    parts = str(label).split('_')
    return parts[-1].strip() if len(parts) > 1 else ""


def load_and_clean_data(file_path):
    if file_path.lower().endswith(".csv"):
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
        header_idx = 0
        for i, line in enumerate(lines):
            if "Location Info" in line and "Last Occurred (ST)" in line:
                header_idx = i
                break
        clean_csv_data = "".join(lines[header_idx:])
        return pd.read_csv(io.StringIO(clean_csv_data), sep=",", engine="python")
    else:
        df = pd.read_excel(file_path, header=None)
        header_idx_list = df[df.apply(lambda r: r.astype(str).str.contains("Location Info", case=False).any(), axis=1)].index
        if not header_idx_list.empty:
            header_idx = header_idx_list[0]
            df.columns = df.iloc[header_idx]
            df = df.iloc[header_idx + 1:].reset_index(drop=True)
        return df


def process_outage_file(file_obj):
    """Processes uploaded CSV/Excel alarm dump files and generates BOTH full & simple excel reports."""
    if file_obj is None:
        return "⚠️ Please upload an Excel or CSV file.", None, None, get_history_table(), generate_handover_summary()

    file_path = file_obj.name
    try:
        df = load_and_clean_data(file_path)
        df.columns = [str(c).strip() for c in df.columns]

        if "Location Info" not in df.columns or "Last Occurred (ST)" not in df.columns:
            return "❌ Error: 'Location Info' or 'Last Occurred (ST)' column not found.", None, None, get_history_table(), generate_handover_summary()

        processed = []
        for _, row in df.iterrows():
            raw_label = str(row["Location Info"]).strip()
            if pd.isna(raw_label) or raw_label.upper() == "NAN" or "VISIBILITY" in raw_label.upper() or not raw_label.upper().startswith("ESS"):
                continue

            processed.append({
                "Service": raw_label,
                "Bandwidth_Text": extract_bw_text(raw_label),
                "Numeric_BW": extract_bw_numeric(raw_label),
                "Client": extract_client_name(raw_label),
                "Category": detect_category(raw_label),
                "UniqueID": extract_unique_id(raw_label),
                "Last Occurred (ST)": row["Last Occurred (ST)"],
            })

        if not processed:
            return "❌ No valid ESS service records found in file.", None, None, get_history_table(), generate_handover_summary()

        df_unsorted = pd.DataFrame(processed).drop_duplicates(subset=["Service"], keep="first").reset_index(drop=True)
        df_sorted = df_unsorted.sort_values(by="Numeric_BW", ascending=False).reset_index(drop=True)

        df_sorted["Time_Parsed"] = pd.to_datetime(df_sorted["Last Occurred (ST)"], errors="coerce")
        valid_times = df_sorted["Time_Parsed"].dropna()
        outlook_time = valid_times.iloc[0].strftime("%Y-%m-%d %I:%M:%S %p") if not valid_times.empty else "Not Found"

        counts = Counter(df_sorted["Category"])
        summary = ", ".join([f"{c}x {s}" for s, c in sorted(counts.items(), key=lambda x: x[1], reverse=True)])

        priority_df = df_sorted[df_sorted["Numeric_BW"] > 250]
        normal_df = df_sorted[df_sorted["Numeric_BW"] <= 250]

        # Generate Console Text Output
        text_lines = []
        text_lines.append("=" * 80)
        text_lines.append(f"{'OUTLOOK READY TEXT':^80}")
        text_lines.append("=" * 80)
        text_lines.append(f"PTN Layer: {len(df_sorted)}x ({summary}) are down")
        text_lines.append("last occured time:")
        text_lines.append(outlook_time)
        text_lines.append("=" * 80 + "\n")

        text_lines.append("Transmission Section\n")
        text_lines.append("Priority Clients")

        for _, row in priority_df.iterrows():
            text_lines.append(row['Service'])

        text_lines.append("\n")

        for _, row in normal_df.iterrows():
            text_lines.append(row['Service'])

        text_lines.append("\nOutage Thread")
        text_lines.append("-" * 50)
        text_lines.append("\n")

        for _, row in df_unsorted.iterrows():
            text_lines.append(row['Service'])

        console_output = "\n".join(text_lines)

        # Append to Shift History Log
        now_str = datetime.now(PKT).strftime("%I:%M %p")
        outage_history.append({
            "id": len(outage_history) + 1,
            "processed_time": now_str,
            "file_name": os.path.basename(file_path),
            "total_links": len(df_sorted),
            "summary": summary,
            "priority_count": len(priority_df),
            "occurred_time": outlook_time,
        })

        file_time = valid_times.iloc[0].strftime("%Y-%m-%d_%I-%M_%p") if not valid_times.empty else datetime.now().strftime("%Y-%m-%d_%I-%M_%p")
        temp_dir = tempfile.gettempdir()

        # 1. Full Categorized Report
        out_filepath = os.path.join(temp_dir, f"Outage_Links_{file_time}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Categorized Links"

        fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        font = Font(bold=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        ws.append(["Part 1: Raw Unsorted Data"])
        ws.cell(row=1, column=1).font = Font(bold=True)
        headers = ["Service", "Bandwidth", "Client", "Category", "UniqueID", "Last Occurred (ST)"]
        ws.append(headers)

        for c in ws[2]:
            c.fill = fill
            c.font = font
            c.border = border

        for _, r in df_unsorted.iterrows():
            ws.append([r["Service"], r["Bandwidth_Text"], r["Client"], r["Category"], r["UniqueID"], r["Last Occurred (ST)"]])

        gap_row = ws.max_row + 3
        ws.cell(row=gap_row, column=1, value="Part 2: Sorted Data for Transmission Rerouting (Descending Bandwidth)")
        ws.cell(row=gap_row, column=1).font = Font(bold=True, italic=True)

        header_row = gap_row + 1
        ws.append(headers)
        for c in ws[header_row]:
            c.fill = fill
            c.font = font
            c.border = border

        for _, r in df_sorted.iterrows():
            bw_val = float(r["Numeric_BW"])
            bw_num = int(bw_val) if bw_val.is_integer() else bw_val
            if bw_num == 0: bw_num = "NA"
            time_fmt = r['Time_Parsed'].strftime("%Y-%m-%d %I:%M:%S %p") if pd.notna(r['Time_Parsed']) else r["Last Occurred (ST)"]
            ws.append([r["Service"], bw_num, r["Client"], r["Category"], r["UniqueID"], time_fmt])

        for col_idx, width in enumerate([75, 15, 35, 15, 25, 25], start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        wb.save(out_filepath)

        # 2. Simple Outage Links Report
        simple_filepath = os.path.join(temp_dir, f"Sample Links.xlsx")
        wb_simple = Workbook()
        ws_simple = wb_simple.active
        ws_simple.title = "Outage Links"

        ws_simple.append(["Links"])
        ws_simple.cell(row=1, column=1).font = Font(bold=True)

        for _, r in df_unsorted.iterrows():
            ws_simple.append([r["Service"]])

        ws_simple.column_dimensions['A'].width = 80
        wb_simple.save(simple_filepath)

        return console_output, out_filepath, simple_filepath, get_history_table(), generate_handover_summary()

    except Exception as e:
        return f"❌ Processing Error: {str(e)}", None, None, get_history_table(), generate_handover_summary()


def get_history_table():
    rows = []
    for item in reversed(outage_history):
        rows.append([
            f"Outage #{item['id']}",
            item['processed_time'],
            item['total_links'],
            item['priority_count'],
            item['summary'],
            item['occurred_time'],
        ])
    return rows


def generate_handover_summary():
    if not outage_history:
        return "No outages processed in the current shift history."

    total_outages = len(outage_history)
    total_links = sum(item['total_links'] for item in outage_history)
    total_priority = sum(item['priority_count'] for item in outage_history)

    lines = [
        "==================================================",
        "📋 SHIFT HANDOVER SUMMARY / DAILY OUTAGE REPORT",
        "==================================================",
        f"• Total Outages Handled Today: {total_outages}",
        f"• Total Down Links Processed: {total_links}",
        f"• High Priority Links (>250M): {total_priority}",
        "--------------------------------------------------",
        "DETAILS BREAKDOWN:",
    ]

    for item in outage_history:
        lines.append(
            f"• Outage #{item['id']} ({item['processed_time']}): "
            f"{item['total_links']} Links Down ({item['summary']}) | "
            f"Occurred: {item['occurred_time']}"
        )

    lines.append("==================================================")
    return "\n".join(lines)


def reset_outage_history():
    global outage_history
    outage_history = []
    return [], "No outages processed in the current shift history."


# ============================================================
# 12. VENDOR ESCALATION MATRIX MANAGER (JSON PERSISTENCE)
# ============================================================
VENDOR_JSON_FILE = "vendors.json"

DEFAULT_VENDOR_MATRIX = {
    "Netsat": [
        {"Level": "Level 1", "Name": "Netsat Support Desk", "Designation": "Help desk / CNOC Netsat", "Time": "1 Hour", "Phone": "021-111638728", "Email": "support.cmpak@netsat.net.pk"},
        {"Level": "Level 1", "Name": "Ali Nadeem", "Designation": "Help desk / CNOC Netsat", "Time": "1 Hour", "Phone": "0309 1112422; 0312-2431968", "Email": "ali.nadeem@netsat.net.pk"},
        {"Level": "Level 2 (Central Region)", "Name": "Zafar Iqbal (CTR)", "Designation": "Regional head (Lahore)", "Time": "3 hours", "Phone": "0301-8114185", "Email": "zafar.iqbal@netsat.net.pk"},
        {"Level": "Level 2 (Central Region)", "Name": "Shoukat Khan (MTR)", "Designation": "Regional head (Multan)", "Time": "3 hours", "Phone": "0302-8271762", "Email": "shoukat.khan@netsat.net.pk"},
        {"Level": "Level 2 (South Region)", "Name": "Farhan (Sindh)", "Designation": "Regional head (Sindh)", "Time": "3 hours", "Phone": "0301-8114182", "Email": "m.farhan@netsat.net.pk"},
        {"Level": "Level 2 (South Region)", "Name": "Syed Yousuf Hussain (Balochistan)", "Designation": "Regional head (Balochistan)", "Time": "3 hours", "Phone": "0300-8259632", "Email": "syed.yousuf@netsat.net.pk"},
        {"Level": "Level 2 (North Region)", "Name": "Asad Imran (North/KPK)", "Designation": "Regional head (North)", "Time": "3 hours", "Phone": "0311-8814294", "Email": "asad.imran@netsat.net.pk"},
        {"Level": "Level 3", "Name": "M. Tahir Qureshi", "Designation": "Operational Support", "Time": "6 hours", "Phone": "0300 8232647", "Email": "m.tahir@netsat.net.pk"},
        {"Level": "Level 3", "Name": "Muhammad Sumair", "Designation": "Operation Head", "Time": "8 hours", "Phone": "0301-8114181", "Email": "sumair@netsat.net.pk"},
        {"Level": "Level 4", "Name": "Syed Mubashir Imam", "Designation": "Senior Operational Head", "Time": "12 hours", "Phone": "0301-8292632", "Email": "mubashir.imam@netsat.net.pk"},
        {"Level": "Level 5", "Name": "Kamran Qaiser", "Designation": "HOD Wireless", "Time": "Urgent/Emergency Maintenance", "Phone": "0301 8114177", "Email": "kamran.qaiser@netsat.net.pk"},
    ],
    "Comstar": [
        {"Level": "Level 1", "Name": "Support Desk", "Designation": "CS", "Time": "10-15min", "Phone": "0333-1312343", "Email": "cs@comstar.com.pk"},
        {"Level": "Level 2", "Name": "Abdul Wajid", "Designation": "TL (South)", "Time": "15-30min", "Phone": "0334-2594529", "Email": "awajid@comstar.com.pk"},
        {"Level": "Level 3", "Name": "Osman Javaid", "Time": "30-45min", "Designation": "Manager Engineering", "Phone": "0336-5479293", "Email": "ojavaid@comstar.com.pk"},
    ],
    "Vision Telecom": [
        {"Level": "Level 1", "Name": "Corporate Helpdesk", "Designation": "Support", "Time": "Immediate", "Phone": "0308-8881418", "Email": "support@visiontelecom.com.pk"},
        {"Level": "Level 2", "Name": "Rizwan Younis", "Designation": "Team Lead", "Time": "30min", "Phone": "0300-0807140", "Email": "rizwan.younis@visiontelecom.com.pk"},
    ],
}


def load_vendors_matrix():
    """Loads vendor matrix dictionary directly from JSON file, initializing defaults if missing."""
    if not os.path.exists(VENDOR_JSON_FILE):
        save_vendors_matrix(DEFAULT_VENDOR_MATRIX)
        return DEFAULT_VENDOR_MATRIX
    try:
        with open(VENDOR_JSON_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if data else DEFAULT_VENDOR_MATRIX
    except Exception:
        return DEFAULT_VENDOR_MATRIX


def save_vendors_matrix(data):
    """Saves updated vendor matrix dictionary directly to vendors.json."""
    try:
        with open(VENDOR_JSON_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving vendors.json: {e}")


vendors_matrix_db = load_vendors_matrix()


def get_vendor_list():
    """Returns dynamic list of all registered vendor names."""
    return list(vendors_matrix_db.keys())


def get_vendor_emails_string(vendor_name):
    """Extracts all emails for a vendor into semicolon-separated string for Outlook."""
    if not vendor_name or vendor_name not in vendors_matrix_db:
        return ""
    emails = []
    for row in vendors_matrix_db[vendor_name]:
        raw_email = row.get("Email", "")
        for e in re.split(r"[;, ]+", raw_email):
            clean_e = e.strip()
            if clean_e and clean_e not in emails:
                emails.append(clean_e)
    return "; ".join(emails)


def render_escalation_matrix_html(vendor_name):
    """Renders HTML matrix table with corporate styling and email CSS selection."""
    if not vendor_name or vendor_name not in vendors_matrix_db:
        return "<p style='color: gray;'>No vendor selected or matrix empty.</p>"

    matrix = vendors_matrix_db[vendor_name]

    html = f"""
    <style>
        .matrix-table,
        .matrix-table th,
        .matrix-table td {{
            -webkit-user-select: none !important;
            -moz-user-select: none !important;
            -ms-user-select: none !important;
            user-select: none !important;
        }}

        .matrix-table td.email-cell,
        .matrix-table td.email-cell a {{
            -webkit-user-select: text !important;
            -moz-user-select: text !important;
            -ms-user-select: text !important;
            user-select: text !important;
            cursor: text !important;
            color: #0000ee !important;
            font-weight: bold;
        }}
    </style>

    <div style="font-family: Arial, sans-serif; max-width: 100%; overflow-x: auto; margin-top: 10px;">
        <table class="matrix-table" style="width: 100%; border-collapse: collapse; border: 1.5px solid #000; font-size: 13px; text-align: left; background:#fff;">
            <thead>
                <tr style="background-color: #7b241c; color: white;">
                    <th colspan="6" style="padding: 10px; font-size: 15px; font-weight: bold; border-bottom: 2px solid #000; text-align:center;">
                        Escalation Matrix - {vendor_name}
                    </th>
                </tr>
                <tr style="background-color: #922b21; color: white; border-bottom: 1.5px solid #000;">
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 18%;">Escalation levels</th>
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 20%;">Name</th>
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 20%;">Designation</th>
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 14%;">Escalation Time</th>
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 14%;">Contact Number</th>
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 14%;">Email address</th>
                </tr>
            </thead>
            <tbody>
    """

    for row in matrix:
        email_val = row.get('Email', '')
        email_html = f'<a href="mailto:{email_val}">{email_val}</a>' if email_val else ''
        html += f"""
        <tr style="border: 1px solid #ccc; background-color: #ffffff; color:#000;">
            <td style="padding: 7px; border: 1px solid #333; font-weight: bold;">{row.get('Level', '')}</td>
            <td style="padding: 7px; border: 1px solid #333;">{row.get('Name', '')}</td>
            <td style="padding: 7px; border: 1px solid #333;">{row.get('Designation', '')}</td>
            <td style="padding: 7px; border: 1px solid #333;">{row.get('Time', '')}</td>
            <td style="padding: 7px; border: 1px solid #333;">{row.get('Phone', '')}</td>
            <td class="email-cell" style="padding: 7px; border: 1px solid #333;">{email_html}</td>
        </tr>
        """

    html += """
            </tbody>
        </table>
    </div>
    """
    return html


def get_matrix_dataframe(vendor_name):
    if not vendor_name or vendor_name not in vendors_matrix_db:
        return []
    return [[r.get("Level", ""), r.get("Name", ""), r.get("Designation", ""), r.get("Time", ""), r.get("Phone", ""), r.get("Email", "")] for r in vendors_matrix_db[vendor_name]]


def add_vendor_name(new_vendor_name):
    """Adds a new vendor entry and syncs to vendors.json."""
    if not new_vendor_name or not new_vendor_name.strip():
        return gr.update(choices=get_vendor_list()), "", "❌ Vendor name is required!"
    v_name = new_vendor_name.strip()
    if v_name not in vendors_matrix_db:
        vendors_matrix_db[v_name] = []
        save_vendors_matrix(vendors_matrix_db)
    return gr.update(choices=get_vendor_list(), value=v_name), get_vendor_emails_string(v_name), f"✅ Vendor '{v_name}' added & saved to vendors.json!"


def add_contact_to_matrix(vendor_name, level, name, designation, time, phone, email):
    """Adds contact person to specified escalation level and auto-saves to vendors.json."""
    if not vendor_name:
        return render_escalation_matrix_html(vendor_name), get_matrix_dataframe(vendor_name), get_vendor_emails_string(vendor_name), "⚠️ Please select a Vendor first!"

    if not name or not email:
        return render_escalation_matrix_html(vendor_name), get_matrix_dataframe(vendor_name), get_vendor_emails_string(vendor_name), "❌ Name and Email are required!"

    vendors_matrix_db[vendor_name].append({
        "Level": level,
        "Name": name.strip(),
        "Designation": designation.strip(),
        "Time": time.strip(),
        "Phone": phone.strip(),
        "Email": email.strip(),
    })

    save_vendors_matrix(vendors_matrix_db)

    return (
        render_escalation_matrix_html(vendor_name),
        get_matrix_dataframe(vendor_name),
        get_vendor_emails_string(vendor_name),
        f"✅ Contact added & saved to vendors.json for {vendor_name}!",
    )


def delete_contact_from_matrix(vendor_name, name_to_delete):
    """Deletes contact person from matrix and auto-saves to vendors.json."""
    if not vendor_name or vendor_name not in vendors_matrix_db:
        return render_escalation_matrix_html(vendor_name), get_matrix_dataframe(vendor_name), get_vendor_emails_string(vendor_name), "⚠️ Vendor not found!"

    vendors_matrix_db[vendor_name] = [r for r in vendors_matrix_db[vendor_name] if r.get("Name") != name_to_delete.strip()]
    save_vendors_matrix(vendors_matrix_db)

    return (
        render_escalation_matrix_html(vendor_name),
        get_matrix_dataframe(vendor_name),
        get_vendor_emails_string(vendor_name),
        f"🗑️ Deleted '{name_to_delete}' & saved changes to vendors.json!",
    )


# ============================================================
# 13. ACTIVE DASHBOARD & MONITORING
# ============================================================
def refresh_dashboard(view_mode):
    records = []

    for label, record in incident_records.items():
        status = record.get("status", "QUEUED")

        if view_mode == "Open Only" and status != "OPEN":
            continue
        if view_mode == "Queued Only" and status != "QUEUED":
            continue
        if view_mode == "Closed Only" and status != "CLOSED":
            continue
        if view_mode == "Queued + Open" and status not in ("QUEUED", "OPEN"):
            continue

        added_time = record.get("added_time")
        reported_time = record.get("reported_time")
        restoration_time = record.get("restoration_time")

        if reported_time:
            age = incident_age_text(
                reported_time,
                added_time=added_time,
                status=status,
            )
        else:
            age = incident_age_text(
                None,
                added_time=added_time,
                status=status,
            )

        duration = (
            calculate_duration(reported_time, restoration_time)
            if reported_time and restoration_time
            else ""
        )

        records.append([
            extract_service_id(label),
            record.get("service_type", ""),
            record.get("issue_type", ""),
            format_dt(added_time) if added_time else "",
            format_dt(reported_time) if reported_time else "Not opened",
            age,
            format_dt(restoration_time) if restoration_time else "",
            duration,
            record.get("ticket", ""),
            record.get("last_stage", ""),
            status,
        ])

    records.sort(
        key=lambda row: row[3] or "",
        reverse=True,
    )
    return records


def check_incident(label):
    label = clean(label)
    if not label:
        return "⚠️ Please enter/select a service label."

    if label not in incident_records:
        return "❌ No runtime complaint record found for this label."

    record = incident_records[label]
    added = record.get("added_time")
    reported = record.get("reported_time")
    restored = record.get("restoration_time")

    duration = (
        calculate_duration(reported, restored)
        if reported and restored
        else "Running" if reported else "Not started"
    )

    return f"""Status: {record.get("status", "QUEUED")}
Service Type: {record.get("service_type", "")}
Issue Type: {record.get("issue_type", "")}
Added to Queue: {format_dt(added) if added else "NA"}
Opening Time: {format_dt(reported) if reported else "Not opened"}
Current Age: {incident_age_text(reported, added, record.get("status", "QUEUED"))}
Closing Time: {format_dt(restored) if restored else "Not closed"}
Total Duration: {duration}
Last Stage: {record.get("last_stage", "")}
Reference Ticket: {record.get("ticket", "")}"""


def reset_incident(label):
    label = clean(label)
    if not label:
        return "⚠️ Please enter/select a service label."

    if label in incident_records:
        del incident_records[label]
        return "✅ Runtime complaint record cleared."

    return "❌ No runtime complaint found for this label."


# ============================================================
# 14. GRADIO UI BUILDER
# ============================================================
def build_app():
    with gr.Blocks(title="Corporate NOC Outage Reporting Console") as app:
        gr.Markdown(
            """
# 🌐 Corporate NOC Outage Reporting Console
Runtime-only console for parallel complaint handling, vendor escalations, troubleshooting, and daily handover reports.
            """
        )

        with gr.Row():
            label = gr.Textbox(
                label="Service Label / Link Name",
                placeholder="ESSClient_DIA_... / SITE-... / link name",
                lines=1,
            )
            common_ticket = gr.Textbox(
                label="Reference Ticket",
                placeholder="Optional",
                lines=1,
            )

        with gr.Tabs():
            # TAB 1: COMPLAINT MANAGER
            with gr.Tab("🗂️ Complaint Manager"):
                gr.Markdown("### Parallel Complaint Queue")

                bulk_complaints = gr.Textbox(
                    label="Add Multiple Complaints",
                    placeholder=(
                        "ESSClient_DPLC_A.A Network_Tandlianwala-FSD_900Mbps_DPLC14722SL1\n"
                        "ESSClient_Central_Turbo_AAA_Brouadband_Faisalabad_2_981Mbps_Turbo14648SL694\n"
                        "ESSClient_SIP_PRI_2_ACE Money Transfer_Kharian_2Mbps_SIP00042"
                    ),
                    lines=8,
                )

                with gr.Row():
                    complaint_default_issue = gr.Dropdown(
                        choices=COMPLAINT_DEFAULT_ISSUES,
                        value="Link is down.",
                        label="Default Complaint Type",
                    )
                    complaint_bulk_ticket = gr.Textbox(
                        label="Reference Ticket (Optional)",
                        placeholder="Applied to newly added complaints",
                    )

                add_complaints_button = gr.Button(
                    "Add Complaints to Runtime Queue",
                    variant="primary",
                )

                with gr.Row():
                    complaint_selector = gr.Dropdown(
                        choices=[],
                        label="Select Runtime Complaint",
                        info="Select a complaint, then load it into the main Service Label field.",
                    )
                    refresh_complaints_button = gr.Button("Refresh Complaint List")

                with gr.Row():
                    load_complaint_button = gr.Button("Load Selected Complaint", variant="primary")
                    remove_complaint_button = gr.Button("Remove Selected Complaint", variant="stop")

                complaint_manager_status = gr.Textbox(
                    label="Complaint Manager Status",
                    lines=5,
                    interactive=False,
                )

                complaint_manager_dashboard = gr.Dataframe(
                    headers=[
                        "Service", "Type", "Issue", "Added", "Opened",
                        "Age", "Closed", "Total Time", "Ticket", "Last Stage", "Status"
                    ],
                    datatype=["str"] * 11,
                    value=[],
                    interactive=False,
                    label="Queued + Open Complaints",
                )

            # TAB 2: OUTAGE ANALYZER
            with gr.Tab("📈 Outage Analyzer"):
                gr.Markdown("### 📊 NOC Alarm & Outage Link Analyzer")
                with gr.Row():
                    with gr.Column():
                        file_input = gr.File(
                            label="Upload Alarm Dump File (CSV / Excel)",
                            file_types=[".csv", ".xlsx", ".xls"],
                        )
                        analyze_button = gr.Button("⚡ Analyze Outage File", variant="primary")
                    with gr.Column():
                        output_file = gr.File(label="📥 Download Categorized Excel Report")
                        simple_links_file = gr.File(label="📄 Download Simple Outage Links Only")

                console_output = gr.Textbox(
                    label="📋 Outlook Ready Summary & Thread Output",
                    lines=14,
                    interactive=False,
                )

                gr.Markdown("---")
                gr.Markdown("### 📜 Shift Outage History & Handover Summary")

                with gr.Row():
                    clear_history_btn = gr.Button("🗑️ Clear Shift History", variant="stop")

                history_table = gr.Dataframe(
                    headers=[
                        "Outage ID", "Processed Time", "Total Links Down",
                        "Priority (>250M)", "Categorized Breakdown", "Last Occurred Time"
                    ],
                    datatype=["str", "str", "number", "number", "str", "str"],
                    value=[],
                    interactive=False,
                    label="Shift Outage Log",
                )

                handover_text = gr.Textbox(
                    label="🗣️ Shift Handover Communication Text (Copy for Next Technical Shift)",
                    lines=8,
                    interactive=False,
                )

                analyze_button.click(
                    fn=process_outage_file,
                    inputs=file_input,
                    outputs=[console_output, output_file, simple_links_file, history_table, handover_text],
                )

                clear_history_btn.click(
                    fn=reset_outage_history,
                    inputs=[],
                    outputs=[history_table, handover_text],
                )

            # TAB 3: VENDOR MATRIX MANAGER
            with gr.Tab("🏪 Vendor Matrix Manager"):
                gr.Markdown("### 📊 Vendor Escalation Matrix Management (JSON Persistence)")

                with gr.Row():
                    with gr.Column(scale=1):
                        gr.Markdown("#### 1️⃣ Select / Add Vendor")
                        vendor_select_dd = gr.Dropdown(
                            choices=get_vendor_list(),
                            value=get_vendor_list()[0] if get_vendor_list() else None,
                            label="Select Vendor",
                        )
                        new_vendor_tb = gr.Textbox(label="Or Add New Vendor Name", placeholder="e.g. Wateen / Cybernet")
                        add_vendor_btn = gr.Button("➕ Add Vendor Name")

                    with gr.Column(scale=2):
                        gr.Markdown("#### 2️⃣ Add Contact Level to Matrix")
                        with gr.Row():
                            level_dd = gr.Dropdown(
                                choices=[
                                    "Level 1",
                                    "Level 2 (Central Region)",
                                    "Level 2 (South Region)",
                                    "Level 2 (North Region)",
                                    "Level 3",
                                    "Level 4",
                                    "Level 5",
                                ],
                                value="Level 1",
                                label="Escalation Level",
                            )
                            c_name_tb = gr.Textbox(label="Contact Name", placeholder="e.g. Ali Nadeem")
                            c_desig_tb = gr.Textbox(label="Designation", placeholder="e.g. Regional Head")

                        with gr.Row():
                            c_time_tb = gr.Textbox(label="Escalation Time", placeholder="e.g. 1 Hour / 3 hours")
                            c_phone_tb = gr.Textbox(label="Contact Phone", placeholder="0300-XXXXXXX")
                            c_email_tb = gr.Textbox(label="Email Address", placeholder="person@vendor.com")

                        add_contact_btn = gr.Button("➕ Add Contact to Matrix", variant="primary")

                matrix_status_msg = gr.Textbox(label="Action Status", interactive=False)

                gr.Markdown("---")
                gr.Markdown("#### 📋 Copy All Emails for Outlook")
                copy_emails_tb = gr.Textbox(
                    label="All Vendor Emails (Copy-paste directly to Outlook To/CC)",
                    value=get_vendor_emails_string(get_vendor_list()[0] if get_vendor_list() else ""),
                    interactive=True,
                )

                gr.Markdown("---")
                gr.Markdown("### 📄 Live Matrix Table Preview (HTML Red Header Format)")

                initial_vendor = get_vendor_list()[0] if get_vendor_list() else ""
                matrix_html_preview = gr.HTML(value=render_escalation_matrix_html(initial_vendor))

                gr.Markdown("---")
                with gr.Row():
                    with gr.Column():
                        gr.Markdown("#### 🗑️ Delete Contact Person")
                        del_name_tb = gr.Textbox(label="Enter Exact Name to Delete", placeholder="e.g. Ali Nadeem")
                        del_contact_btn = gr.Button("🗑️ Remove Person", variant="stop")

                    with gr.Column():
                        gr.Markdown("#### 📋 Dataframe View")
                        matrix_df_view = gr.Dataframe(
                            headers=["Level", "Name", "Designation", "Time", "Phone", "Email"],
                            value=get_matrix_dataframe(initial_vendor),
                            interactive=False,
                        )

                vendor_select_dd.change(
                    fn=lambda v: (render_escalation_matrix_html(v), get_matrix_dataframe(v), get_vendor_emails_string(v)),
                    inputs=vendor_select_dd,
                    outputs=[matrix_html_preview, matrix_df_view, copy_emails_tb],
                )

                add_vendor_btn.click(
                    fn=add_vendor_name,
                    inputs=new_vendor_tb,
                    outputs=[vendor_select_dd, copy_emails_tb, matrix_status_msg],
                )

                add_contact_btn.click(
                    fn=add_contact_to_matrix,
                    inputs=[vendor_select_dd, level_dd, c_name_tb, c_desig_tb, c_time_tb, c_phone_tb, c_email_tb],
                    outputs=[matrix_html_preview, matrix_df_view, copy_emails_tb, matrix_status_msg],
                )

                del_contact_btn.click(
                    fn=delete_contact_from_matrix,
                    inputs=[vendor_select_dd, del_name_tb],
                    outputs=[matrix_html_preview, matrix_df_view, copy_emails_tb, matrix_status_msg],
                )

            # TAB 4: OPENING
            with gr.Tab("🚨 Opening"):
                gr.Markdown("### Open / Start a Complaint")

                with gr.Row():
                    opening_complaint_selector = gr.Dropdown(
                        choices=[],
                        label="Select Complaint to Open",
                    )
                    opening_refresh_complaints = gr.Button("Refresh Opening Complaint List")

                opening_selected_details = gr.Textbox(
                    label="Selected Complaint Details",
                    lines=3,
                    interactive=False,
                )

                opening_issue = gr.Dropdown(
                    choices=OPENING_ISSUES,
                    value="Link is down.",
                    label="Issue Type",
                )
                opening_custom_issue = gr.Textbox(
                    label="Custom Issue",
                    placeholder="Used only when Other / Manual is selected",
                )
                opening_priority = gr.Dropdown(
                    choices=["Normal", "Urgent", "Critical"],
                    value="Normal",
                    label="Priority",
                )

                gr.Markdown("### 🕐 Opening Clock")
                opening_time_mode = gr.Radio(
                    choices=[AUTO_TIME_MODE, MANUAL_TIME_MODE],
                    value=AUTO_TIME_MODE,
                    label="Opening Time Mode",
                )

                with gr.Row():
                    opening_date = gr.DateTime(
                        value=None,
                        include_time=False,
                        type="string",
                        label="Opening Date",
                        interactive=True,
                    )
                    opening_hour = gr.Dropdown(
                        choices=HOUR_CHOICES,
                        value=None,
                        label="Hour",
                    )
                    opening_minute = gr.Dropdown(
                        choices=MINUTE_CHOICES,
                        value=None,
                        label="Minute",
                    )
                    opening_ampm = gr.Radio(
                        choices=AMPM_CHOICES,
                        value=None,
                        label="AM / PM",
                    )

                opening_set_current = gr.Button("🕐 Set Manual Picker to Current Pakistan Time")
                opening_clock_preview = gr.HTML(
                    value=clock_card_html(),
                    label="Opening Clock Preview",
                )

                opening_button = gr.Button("Generate Opening Email", variant="primary")
                opening_subject = gr.Textbox(label="Subject", interactive=True)
                opening_body = gr.Textbox(label="Opening Email", lines=12, interactive=True)
                opening_runtime = gr.Textbox(label="Runtime Status", interactive=False)

            # TAB 5: CUSTOMER END / FINDINGS
            with gr.Tab("👤 Customer End / Findings"):
                customer_issue_summary = gr.Textbox(
                    label="Issue Summary",
                    value="Reported service issue",
                )
                customer_findings = gr.CheckboxGroup(
                    choices=FINDING_OPTIONS,
                    label="Select Findings",
                )
                customer_action = gr.Dropdown(
                    choices=CUSTOMER_ACTIONS,
                    value="Verify last-mile media",
                    label="Requested Customer Action",
                )
                customer_priority = gr.Dropdown(
                    choices=["Normal", "Follow-up", "Urgent"],
                    value="Normal",
                    label="Priority",
                )

                customer_end_button = gr.Button("Generate Customer-End Response", variant="primary")
                customer_end_subject = gr.Textbox(label="Subject", interactive=True)
                customer_end_body = gr.Textbox(label="Generated Email", lines=12, interactive=True)
                customer_policy = gr.Textbox(label="Internal Policy / Warning", lines=2, interactive=False)

            # TAB 6: VENDOR / INTERNAL ESCALATION (PLACED NEXT TO CUSTOMER END)
            with gr.Tab("📨 Vendor / Internal Escalation"):
                gr.Markdown("### 📤 Vendor & Internal Team Escalation Console")

                with gr.Row():
                    escalation_format_radio = gr.Radio(
                        choices=ESCALATION_FORMAT_OPTIONS,
                        value="Option A (Standard Matrix Format)",
                        label="Email & Subject Format Mode",
                    )

                with gr.Row():
                    escalation_target = gr.Dropdown(
                        choices=ESCALATION_TARGETS,
                        value="Netsat",
                        label="Escalate To (Vendor)",
                    )
                    custom_escalation_target = gr.Textbox(
                        label="Custom Vendor Name",
                        placeholder="Used only when 'Other' is selected",
                    )
                    rcbs_group_dd = gr.Dropdown(
                        choices=["None", "RCBS North", "RCBS Central", "RCBS South"],
                        value="RCBS Central",
                        label="CC (RCBS Region Group)",
                    )

                # Editable To & CC inputs
                with gr.Row():
                    escalation_to_tb = gr.Textbox(
                        label="To: Recipients (Auto-filled L1 + L2 | Add/Edit Manual Addresses Here)",
                        value=get_vendor_l1_l2_emails("Netsat"),
                        placeholder="support.cmpak@netsat.net.pk; engineer@vendor.com",
                        lines=2,
                        interactive=True,
                    )
                    escalation_cc_tb = gr.Textbox(
                        label="Cc: Recipients (Auto-filled RCBS Group | Add/Edit Manual CC Here)",
                        value=RCBS_GROUPS.get("RCBS Central", ""),
                        placeholder="group.rcbs.central@gmail.com; supervisor@company.com",
                        lines=2,
                        interactive=True,
                    )

                # Option C Specific Custom Subject input
                with gr.Row():
                    custom_c_subject = gr.Textbox(
                        label="Custom Subject (Used primarily for Option C)",
                        placeholder="e.g. Lahore ZONG PRI DOWN vlan 345",
                        lines=1,
                    )

                with gr.Row():
                    escalation_level = gr.Dropdown(
                        choices=ESCALATION_LEVELS,
                        value="Initial engagement",
                        label="Escalation Level",
                    )
                    escalation_priority = gr.Dropdown(
                        choices=["Normal", "Follow-up", "Urgent", "Critical"],
                        value="Urgent",
                        label="Priority Tag",
                    )

                escalation_findings = gr.CheckboxGroup(
                    choices=FINDING_OPTIONS,
                    label="Current Findings (Optional)",
                )

                escalation_action = gr.Textbox(
                    label="Required Action (Optional)",
                    placeholder="e.g. Kindly dispatch a team to the site.",
                    lines=2,
                )

                escalation_button = gr.Button("🚀 Generate Escalation Email & Outlook File", variant="primary")

                escalation_subject = gr.Textbox(label="Generated Subject", interactive=True)
                escalation_body = gr.Textbox(label="Generated Email Body", lines=12, interactive=True)

                outlook_file_output = gr.File(label="📥 Download Ready Outlook File (.eml)")
                matrix_table_output = gr.HTML(label="Vendor Escalation Matrix Live Preview")

                # Dynamic Recipient Pre-Filling Events
                escalation_target.change(
                    fn=lambda v: get_vendor_l1_l2_emails(v),
                    inputs=escalation_target,
                    outputs=escalation_to_tb,
                )

                rcbs_group_dd.change(
                    fn=lambda g: RCBS_GROUPS.get(g, ""),
                    inputs=rcbs_group_dd,
                    outputs=escalation_cc_tb,
                )

            # TAB 7: CLOSURE / RFO
            with gr.Tab("✅ Closure / RFO"):
                gr.Markdown("### Close an Active Complaint")

                with gr.Row():
                    closure_complaint_selector = gr.Dropdown(
                        choices=[],
                        label="Select OPEN Complaint to Close",
                    )
                    closure_refresh_complaints = gr.Button("Refresh Open Complaint List")

                closure_selected_details = gr.Textbox(
                    label="Selected Complaint Details",
                    lines=3,
                    interactive=False,
                )

                with gr.Row():
                    issue_found_at = gr.Dropdown(
                        choices=ISSUE_FOUND_AT,
                        value="Customer",
                        label="Issue Found At",
                    )
                    custom_issue_found_at = gr.Textbox(
                        label="Custom Issue Found At",
                        placeholder="Used only for Other",
                    )

                root_cause = gr.Dropdown(
                    choices=RCA_OPTIONS,
                    value="No Issue Observed",
                    label="Root Cause",
                )
                custom_root_cause = gr.Textbox(
                    label="Custom Root Cause",
                    placeholder="Used only for Other / Manual",
                )

                corrective_action = gr.Dropdown(
                    choices=[
                        "Auto from Root Cause", "No activity performed at our end",
                        "Fiber restored", "Configuration rectified", "Link rerouted",
                        "Power restored", "Wireless parameters optimized",
                        "Optical parameters normalized", "Faulty equipment replaced",
                        "Port configuration corrected", "Issue resolved by upstream provider",
                        "Customer end issue rectified", "Link stabilized after troubleshooting",
                        "Other / Manual",
                    ],
                    value="Auto from Root Cause",
                    label="Corrective Action",
                )
                custom_corrective_action = gr.Textbox(
                    label="Custom Corrective Action",
                    placeholder="Used only for Other / Manual",
                )

                final_status = gr.Dropdown(
                    choices=[
                        "Link is up and stable.", "Service is up and working normally.",
                        "Wireless segment is up and normal.", "Link parameters are normal.",
                        "Service has been restored successfully.", "No abnormality is currently observed at our end.",
                    ],
                    value="Service is up and working normally.",
                    label="Final Status",
                )
                closure_priority = gr.Dropdown(
                    choices=["Normal", "Urgent"],
                    value="Normal",
                    label="Priority",
                )

                gr.Markdown("### 🕐 Closing / Restoration Clock")
                closure_time_mode = gr.Radio(
                    choices=[AUTO_TIME_MODE, MANUAL_TIME_MODE],
                    value=AUTO_TIME_MODE,
                    label="Closing Time Mode",
                )

                with gr.Row():
                    closure_date = gr.DateTime(
                        value=None,
                        include_time=False,
                        type="string",
                        label="Restoration Date",
                        interactive=True,
                    )
                    closure_hour = gr.Dropdown(
                        choices=HOUR_CHOICES,
                        value=None,
                        label="Hour",
                    )
                    closure_minute = gr.Dropdown(
                        choices=MINUTE_CHOICES,
                        value=None,
                        label="Minute",
                    )
                    closure_ampm = gr.Radio(
                        choices=AMPM_CHOICES,
                        value=None,
                        label="AM / PM",
                    )

                closure_set_current = gr.Button("🕐 Set Manual Picker to Current Pakistan Time")
                closure_clock_preview = gr.HTML(
                    value=clock_card_html(),
                    label="Closing Clock Preview",
                )

                closure_button = gr.Button("Generate Closure Email", variant="primary")
                closure_subject = gr.Textbox(label="Subject", interactive=True)
                closure_body = gr.Textbox(label="Generated Closure Email", lines=14, interactive=True)
                closure_runtime = gr.Textbox(label="Runtime Status", interactive=False)

            # TAB 8: TROUBLESHOOTING / STATS
            with gr.Tab("🧪 Stats / Troubleshooting"):
                stats_scenario = gr.Dropdown(
                    choices=STATS_SCENARIOS,
                    value="Packet Loss / High Latency",
                    label="Reported Scenario",
                )
                stats_audience = gr.Dropdown(
                    choices=["Customer", "Team / Vendor"],
                    value="Customer",
                    label="Audience",
                )
                stats_context = gr.Textbox(
                    label="Optional Additional Context",
                    placeholder="e.g. Please perform testing from the same affected IP pool.",
                    lines=2,
                )
                stats_button = gr.Button("Generate Required-Stats Email", variant="primary")
                stats_subject = gr.Textbox(label="Subject", interactive=True)
                stats_body = gr.Textbox(label="Generated Email", lines=14, interactive=True)
                stats_policy = gr.Textbox(label="Internal Policy / Warning", lines=2, interactive=False)

            # TAB 9: PROGRESS UPDATE
            with gr.Tab("🔄 Progress Update"):
                progress_status = gr.Dropdown(
                    choices=PROGRESS_OPTIONS,
                    value="Concerned team engaged",
                    label="Current Progress",
                )
                with gr.Row():
                    progress_ettr = gr.Dropdown(
                        choices=[
                            "Awaited", "30 Minutes", "1 Hour", "2 Hours",
                            "4 Hours", "Not Available", "Not Applicable", "Custom",
                        ],
                        value="Awaited",
                        label="ETTR",
                    )
                    progress_custom_ettr = gr.Textbox(label="Custom ETTR", placeholder="Optional")
                    progress_priority = gr.Dropdown(
                        choices=["Normal", "Follow-up", "Urgent", "Critical"],
                        value="Normal",
                        label="Priority",
                    )
                    progress_audience = gr.Dropdown(
                        choices=["Customer", "Team / Vendor"],
                        value="Customer",
                        label="Audience",
                    )

                progress_note = gr.Textbox(label="Additional Ground Update", placeholder="Optional", lines=2)
                progress_button = gr.Button("Generate Progress Update", variant="primary")
                progress_subject = gr.Textbox(label="Subject", interactive=True)
                progress_body = gr.Textbox(label="Generated Email", lines=12, interactive=True)

            # TAB 10: ACTIVE COMPLAINTS DASHBOARD
            with gr.Tab("📋 Active Complaints / Response Age"):
                with gr.Row():
                    dashboard_view = gr.Dropdown(
                        choices=[
                            "Queued + Open", "Open Only", "Queued Only",
                            "Closed Only", "All Runtime Complaints",
                        ],
                        value="Queued + Open",
                        label="View",
                    )
                    dashboard_complaint_selector = gr.Dropdown(
                        choices=[],
                        label="Select Complaint from Runtime",
                    )

                with gr.Row():
                    dashboard_refresh = gr.Button("Refresh Dashboard", variant="primary")
                    dashboard_selector_refresh = gr.Button("Refresh Complaint Selector")
                    dashboard_load_button = gr.Button("Load Selected Complaint")

                dashboard = gr.Dataframe(
                    headers=[
                        "Service", "Type", "Issue", "Added", "Opened",
                        "Age", "Closed", "Total Time", "Ticket", "Last Stage", "Status"
                    ],
                    datatype=["str"] * 11,
                    value=[],
                    interactive=False,
                    label="Runtime Complaints",
                )

                check_button = gr.Button("Check Current Label")
                incident_status = gr.Textbox(
                    label="Current Label Runtime Details",
                    lines=8,
                    interactive=False,
                )
                reset_button = gr.Button("Clear Current Label from Runtime")

        # ====================================================
        # GLOBAL EVENT BINDINGS
        # ====================================================
        add_complaints_event = add_complaints_button.click(
            fn=register_complaints,
            inputs=[bulk_complaints, complaint_default_issue, complaint_bulk_ticket],
            outputs=[complaint_selector, complaint_manager_status, complaint_manager_dashboard],
        )

        add_complaints_event.then(
            fn=refresh_opening_complaint_selector,
            inputs=[],
            outputs=opening_complaint_selector,
        ).then(
            fn=refresh_closure_complaint_selector,
            inputs=[],
            outputs=closure_complaint_selector,
        ).then(
            fn=refresh_complaint_selector,
            inputs=[],
            outputs=dashboard_complaint_selector,
        )

        refresh_complaints_button.click(
            fn=refresh_complaint_selector,
            inputs=[],
            outputs=complaint_selector,
        )

        load_complaint_button.click(
            fn=load_selected_complaint,
            inputs=complaint_selector,
            outputs=[label, common_ticket, complaint_manager_status],
        )

        opening_refresh_complaints.click(
            fn=refresh_opening_complaint_selector,
            inputs=[],
            outputs=opening_complaint_selector,
        )

        opening_complaint_selector.change(
            fn=load_selected_complaint,
            inputs=opening_complaint_selector,
            outputs=[label, common_ticket, opening_selected_details],
        )

        closure_refresh_complaints.click(
            fn=refresh_closure_complaint_selector,
            inputs=[],
            outputs=closure_complaint_selector,
        )

        closure_complaint_selector.change(
            fn=load_selected_complaint,
            inputs=closure_complaint_selector,
            outputs=[label, common_ticket, closure_selected_details],
        )

        remove_complaint_button.click(
            fn=remove_selected_complaint,
            inputs=complaint_selector,
            outputs=[complaint_selector, complaint_manager_status, complaint_manager_dashboard],
        )

        opening_set_current.click(
            fn=current_clock_picker_values,
            inputs=[],
            outputs=[
                opening_time_mode, opening_date, opening_hour,
                opening_minute, opening_ampm, opening_clock_preview,
            ],
        )

        for _comp in [opening_date, opening_hour, opening_minute, opening_ampm]:
            _comp.change(
                fn=update_clock_preview,
                inputs=[opening_date, opening_hour, opening_minute, opening_ampm],
                outputs=opening_clock_preview,
            )

        closure_set_current.click(
            fn=current_clock_picker_values,
            inputs=[],
            outputs=[
                closure_time_mode, closure_date, closure_hour,
                closure_minute, closure_ampm, closure_clock_preview,
            ],
        )

        for _comp in [closure_date, closure_hour, closure_minute, closure_ampm]:
            _comp.change(
                fn=update_clock_preview,
                inputs=[closure_date, closure_hour, closure_minute, closure_ampm],
                outputs=closure_clock_preview,
            )

        opening_event = opening_button.click(
            fn=generate_opening_email,
            inputs=[
                opening_complaint_selector, label, opening_issue, opening_custom_issue,
                common_ticket, opening_priority, opening_time_mode, opening_date,
                opening_hour, opening_minute, opening_ampm,
            ],
            outputs=[label, opening_subject, opening_body, opening_runtime],
        )

        opening_event.then(
            fn=refresh_opening_complaint_selector,
            inputs=[],
            outputs=opening_complaint_selector,
        ).then(
            fn=refresh_closure_complaint_selector,
            inputs=[],
            outputs=closure_complaint_selector,
        )

        customer_end_button.click(
            fn=generate_customer_end_email,
            inputs=[
                label, customer_issue_summary, customer_findings,
                customer_action, customer_priority, common_ticket,
            ],
            outputs=[customer_end_subject, customer_end_body, customer_policy],
        )

        # Escalation Event Binding with Manual To/CC inputs
        escalation_button.click(
            fn=generate_escalation_email,
            inputs=[
                label, escalation_target, escalation_level, escalation_format_radio,
                rcbs_group_dd, escalation_to_tb, escalation_cc_tb, custom_c_subject,
                common_ticket, escalation_findings, escalation_action,
                escalation_priority, custom_escalation_target,
            ],
            outputs=[
                escalation_subject, escalation_to_tb, escalation_cc_tb,
                escalation_body, outlook_file_output, matrix_table_output,
            ],
        )

        stats_button.click(
            fn=generate_stats_request,
            inputs=[label, stats_scenario, stats_audience, stats_context],
            outputs=[stats_subject, stats_body, stats_policy],
        )

        progress_button.click(
            fn=generate_progress_email,
            inputs=[
                label, progress_status, progress_ettr, progress_custom_ettr,
                progress_note, progress_priority, progress_audience,
            ],
            outputs=[progress_subject, progress_body],
        )

        closure_event = closure_button.click(
            fn=generate_closure_email,
            inputs=[
                closure_complaint_selector, label, issue_found_at, custom_issue_found_at,
                root_cause, custom_root_cause, corrective_action, custom_corrective_action,
                final_status, closure_priority, closure_time_mode, closure_date,
                closure_hour, closure_minute, closure_ampm,
            ],
            outputs=[label, closure_subject, closure_body, closure_runtime],
        )

        closure_event.then(
            fn=refresh_closure_complaint_selector,
            inputs=[],
            outputs=closure_complaint_selector,
        ).then(
            fn=refresh_opening_complaint_selector,
            inputs=[],
            outputs=opening_complaint_selector,
        ).then(
            fn=refresh_complaint_selector,
            inputs=[],
            outputs=dashboard_complaint_selector,
        )

        dashboard_refresh.click(
            fn=refresh_dashboard,
            inputs=dashboard_view,
            outputs=dashboard,
        )

        dashboard_selector_refresh.click(
            fn=refresh_complaint_selector,
            inputs=[],
            outputs=dashboard_complaint_selector,
        )

        dashboard_load_button.click(
            fn=load_selected_complaint,
            inputs=dashboard_complaint_selector,
            outputs=[label, common_ticket, incident_status],
        )

        check_button.click(
            fn=check_incident,
            inputs=label,
            outputs=incident_status,
        )

        reset_button.click(
            fn=reset_incident,
            inputs=label,
            outputs=incident_status,
        )

    return app


# ============================================================
# 15. APPLICATION LAUNCH (COLAB READY & PORTLESS)
# ============================================================
app = build_app()
if __name__ == "__main__":
    gr.close_all()
    app.launch(
        share=True,
        auth=[
            ("nasir", "123"),
        ],
        auth_message="🔒 Corporate NOC Response Console - Authorized Personnel Only",
    )# ============================================================
# 1. IMPORTS & GLOBAL CONFIGURATION
# ============================================================
import os
import re
import io
import json
import math
import tempfile
import warnings
from datetime import datetime
from collections import Counter
from zoneinfo import ZoneInfo
from email.message import EmailMessage

import pandas as pd
import gradio as gr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

warnings.filterwarnings("ignore")

# Timezone Configuration (Pakistan Standard Time - PKT)
PKT = ZoneInfo("Asia/Karachi")

# Runtime Memory Incident Tracking Database
incident_records = {}

# RCBS Group Email Mapping for CC
RCBS_GROUPS = {
    "None": "",
    "RCBS North": "group.rcbs.north@gmail.com",
    "RCBS Central": "group.rcbs.central@gmail.com",
    "RCBS South": "group.rcbs.south@gmail.com",
}


# ============================================================
# 2. SHARED HELPERS & TIME FUNCTIONS
# ============================================================
def clean(value):
    """Clean and strip input string safely."""
    return (value or "").strip()


def get_current_time():
    """Returns current datetime in Asia/Karachi timezone."""
    return datetime.now(PKT)


def format_dt(value):
    """Formats datetime object to standard corporate representation."""
    if not value:
        return "NA"
    return value.strftime("%d-%m-%Y %I:%M %p")


def detect_service_type(text):
    """Detects standard telecom service type from label or complaint string."""
    value = clean(text).upper()

    checks = [
        ("BGP", "BGP DIA"),
        ("MPLS", "MPLS"),
        ("DPLC", "DPLC"),
        ("TURBONET", "Turbonet"),
        ("TURBO", "Turbonet"),
        ("SIP PRI", "SIP PRI"),
        ("SIP_PRI", "SIP PRI"),
        ("IPLC", "IPLC"),
        ("DARKCORE", "Darkcore Fiber"),
        ("DARK CORE", "Darkcore Fiber"),
        ("M2M", "M2M"),
        ("PRI", "PRI"),
        ("SIP", "SIP"),
        ("DIA", "DIA"),
    ]

    for token, service in checks:
        if token in value:
            return service

    return "Corporate"


def extract_service_id(label):
    """Extracts short unique service link identifier for concise subjects."""
    value = clean(label)
    if not value:
        return "Service"

    patterns = [
        r"\bDPLC\d+SL\d+\b",
        r"\bDIA\d+SL\d+\b",
        r"\bTurbo\d+SL\d+\b",
        r"\bLNK[A-Z0-9]+\b",
        r"\bMPLS[A-Z0-9-]*\b",
        r"\bPRI[A-Z0-9-]*\b",
        r"\bSIP[A-Z0-9-]*\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, value, flags=re.IGNORECASE)
        if match:
            return match.group(0)

    if "_" in value:
        candidate = value.split("_")[-1].strip()
        if candidate:
            return candidate[:70]

    return value[:70]


def priority_prefix(priority):
    """Returns email subject prefix tag according to escalation priority."""
    mapping = {
        "Normal": "",
        "Follow-up": "FOLLOW-UP | ",
        "Urgent": "URGENT | ",
        "Critical": "CRITICAL | ",
    }
    return mapping.get(priority, "")


def make_subject(label, topic, priority="Normal", ticket=""):
    """Standardized email subject generator."""
    service_id = extract_service_id(label)
    ticket = clean(ticket)
    ticket_part = f" | {ticket}" if ticket else ""
    return f"{priority_prefix(priority)}{topic} | {service_id}{ticket_part}"


def salutation(audience):
    """Returns appropriate professional salutation based on target audience."""
    if audience == "Customer":
        return "Dear Customer,"
    return "Dear Team,"


def calculate_duration(start_time, end_time):
    """Calculates formatted total outage duration between two datetime objects."""
    if not start_time or not end_time:
        return "NA"

    total_minutes = max(0, int((end_time - start_time).total_seconds() // 60))
    days, remainder = divmod(total_minutes, 1440)
    hours, minutes = divmod(remainder, 60)

    if days:
        return f"{days} Day(s) {hours} Hour(s) {minutes} Minute(s)"
    if hours:
        return f"{hours} Hour(s) {minutes} Minute(s)"
    return f"{minutes} Minute(s)"


def elapsed_text(start_time, end_time=None):
    """Calculates concise running age string for active incidents."""
    if not start_time:
        return "Not started"

    end_time = end_time or get_current_time()
    minutes = max(0, int((end_time - start_time).total_seconds() // 60))

    if minutes >= 1440:
        days, rem = divmod(minutes, 1440)
        hours, mins = divmod(rem, 60)
        return f"{days}d {hours}h {mins}m"
    if minutes >= 60:
        hours, mins = divmod(minutes, 60)
        return f"{hours}h {mins}m"
    return f"{minutes}m"


def incident_age_text(reported_time, added_time=None, status="OPEN"):
    """Generates visual status indicators based on active response age."""
    base_time = reported_time or added_time
    if not base_time:
        return "—"

    if status == "CLOSED":
        return "Closed"

    minutes = max(0, int((get_current_time() - base_time).total_seconds() // 60))

    if minutes < 5:
        marker = "🟢"
    elif minutes < 15:
        marker = "🟡"
    elif minutes < 30:
        marker = "🟠"
    else:
        marker = "🔴"

    prefix = "Queued " if not reported_time else ""
    return f"{marker} {prefix}{elapsed_text(base_time)}"


TIME_MODES = [
    "Automatic - Current Pakistan Time",
    "Manual Date/Time",
]


def parse_datetime_input(value):
    """Parses various date-time format strings into a PKT-aware datetime object."""
    value = clean(value)
    if not value:
        raise ValueError("Manual date/time is empty.")

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %I:%M:%S %p",
        "%Y-%m-%d %I:%M %p",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %I:%M:%S %p",
        "%d-%m-%Y %I:%M %p",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %I:%M:%S %p",
        "%d/%m/%Y %I:%M %p",
    ]

    for fmt in formats:
        try:
            parsed = datetime.strptime(value, fmt)
            return parsed.replace(tzinfo=PKT)
        except ValueError:
            continue

    raise ValueError(
        "Invalid date/time. Use YYYY-MM-DD HH:MM, DD-MM-YYYY HH:MM, or include AM/PM."
    )


def update_incident_stage(label, stage):
    """Updates the last stage marker of an existing active incident."""
    label = clean(label)
    if label in incident_records:
        incident_records[label]["last_stage"] = stage


def format_ettr(ettr_choice, custom_ettr):
    """Formats standardized Expected Time to Restore (ETTR) statements."""
    if ettr_choice == "Awaited":
        return "ETTR will be shared once received from the concerned team."
    if ettr_choice == "Not Available":
        return "Currently, no confirmed ETTR is available. Further updates will be shared accordingly."
    if ettr_choice == "Not Applicable":
        return ""
    if ettr_choice == "Custom":
        value = clean(custom_ettr)
        return (
            f"The tentative ETTR shared by the concerned team is {value}."
            if value
            else "ETTR will be shared once received from the concerned team."
        )
    return f"The tentative ETTR shared by the concerned team is {ettr_choice}."


def multiline_links(value):
    """Splits multiline input text into clean lists of individual link labels."""
    entries = []
    for raw in clean(value).splitlines():
        item = raw.strip(" \t-•")
        if item:
            entries.append(item)
    return entries


# ============================================================
# 3. COMPLAINT MANAGER ENGINE
# ============================================================
COMPLAINT_DEFAULT_ISSUES = [
    "Link is down.",
    "Service degradation.",
    "Intermittent connectivity.",
    "High latency observed.",
    "Packet loss observed.",
    "Slow browsing issue.",
    "Bandwidth issue.",
    "Website / application issue.",
    "Other / To be classified",
]


def get_complaint_labels(include_closed=True):
    """Retrieves list of runtime complaint labels sorted by addition time."""
    labels = []
    for label, record in incident_records.items():
        if not include_closed and record.get("status") == "CLOSED":
            continue
        labels.append(label)

    labels.sort(
        key=lambda x: incident_records[x].get("added_time") or get_current_time(),
        reverse=True,
    )
    return labels


def register_complaints(raw_labels, default_issue, ticket):
    """Registers multiple complaints into runtime memory queue simultaneously."""
    entries = multiline_links(raw_labels)
    if not entries:
        return (
            gr.Dropdown(choices=get_complaint_labels(), value=None),
            "⚠️ Paste one or more complaints/service labels, one per line.",
            refresh_dashboard("Queued + Open"),
        )

    now = get_current_time()
    added = []
    skipped = []
    restarted = []

    for label in entries:
        existing = incident_records.get(label)

        if existing and existing.get("status") in ("QUEUED", "OPEN"):
            skipped.append(label)
            continue

        if existing and existing.get("status") == "CLOSED":
            restarted.append(label)

        incident_records[label] = {
            "added_time": now,
            "reported_time": None,
            "restoration_time": None,
            "service_type": detect_service_type(label),
            "issue_type": clean(default_issue) or "Other / To be classified",
            "ticket": clean(ticket),
            "status": "QUEUED",
            "last_stage": "Complaint added - awaiting opening",
        }
        added.append(label)

    choices = get_complaint_labels()
    selected = added[0] if added else (choices[0] if choices else None)

    parts = []
    if added:
        parts.append(f"✅ Added {len(added)} complaint(s) to runtime queue.")
    if restarted:
        parts.append(
            f"♻️ {len(restarted)} previously closed label(s) started as new queued complaint(s)."
        )
    if skipped:
        parts.append(
            f"ℹ️ Skipped {len(skipped)} duplicate complaint(s) already QUEUED/OPEN."
        )

    return (
        gr.Dropdown(choices=choices, value=selected),
        "\n".join(parts),
        refresh_dashboard("Queued + Open"),
    )


def load_selected_complaint(selected_label):
    """Loads complaint details into active editing state."""
    label = clean(selected_label)
    if not label or label not in incident_records:
        return "", "", "⚠️ Please select a runtime complaint."

    record = incident_records[label]
    status = record.get("status", "")
    opened = (
        format_dt(record["reported_time"])
        if record.get("reported_time")
        else "Not opened yet"
    )
    closed = (
        format_dt(record["restoration_time"])
        if record.get("restoration_time")
        else "Not closed"
    )

    summary = f"""✅ Complaint loaded.
Status: {status}
Service Type: {record.get("service_type", "")}
Issue: {record.get("issue_type", "")}
Opening Time: {opened}
Closing Time: {closed}
Ticket: {record.get("ticket", "")}"""

    return label, record.get("ticket", ""), summary


def remove_selected_complaint(selected_label):
    """Deletes complaint record from active memory."""
    label = clean(selected_label)
    if not label or label not in incident_records:
        return (
            gr.Dropdown(choices=get_complaint_labels(), value=None),
            "⚠️ Please select a valid runtime complaint.",
            refresh_dashboard("Queued + Open"),
        )

    del incident_records[label]
    choices = get_complaint_labels()
    selected = choices[0] if choices else None

    return (
        gr.Dropdown(choices=choices, value=selected),
        "✅ Selected complaint removed from runtime memory.",
        refresh_dashboard("Queued + Open"),
    )


def refresh_complaint_selector():
    """Refreshes dropdown list choices of complaints."""
    choices = get_complaint_labels()
    return gr.Dropdown(
        choices=choices,
        value=choices[0] if choices else None,
    )


def get_complaint_labels_by_status(statuses=None):
    """Filters complaint labels by status (QUEUED, OPEN, CLOSED)."""
    labels = []
    allowed = set(statuses) if statuses else None

    for complaint_label, record in incident_records.items():
        status = record.get("status", "QUEUED")
        if allowed is not None and status not in allowed:
            continue
        labels.append(complaint_label)

    labels.sort(
        key=lambda x: incident_records[x].get("added_time") or get_current_time(),
        reverse=True,
    )
    return labels


def refresh_opening_complaint_selector():
    """Refreshes list choices for Opening Email tab."""
    choices = get_complaint_labels_by_status({"QUEUED", "OPEN"})
    return gr.Dropdown(
        choices=choices,
        value=choices[0] if choices else None,
    )


def refresh_closure_complaint_selector():
    """Refreshes list choices for Closure Email tab."""
    choices = get_complaint_labels_by_status({"OPEN"})
    return gr.Dropdown(
        choices=choices,
        value=choices[0] if choices else None,
    )


def resolve_complaint_label(selected_label, fallback_label):
    """Resolves label between selected complaint dropdown or textbox fallback."""
    return clean(selected_label) or clean(fallback_label)


# ============================================================
# 4. CLICK-ONLY DATE / CLOCK PICKER HELPERS
# ============================================================
AUTO_TIME_MODE = "Automatic - Current Pakistan Time"
MANUAL_TIME_MODE = "Manual - Mouse Clock Picker"

HOUR_CHOICES = [f"{i:02d}" for i in range(1, 13)]
MINUTE_CHOICES = [f"{i:02d}" for i in range(60)]
AMPM_CHOICES = ["AM", "PM"]


def _picker_date_to_date(value):
    """Helper to convert string/datetime date pickers into python date objects."""
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.astimezone(PKT).date() if value.tzinfo else value.date()

    if isinstance(value, (int, float)):
        return datetime.fromtimestamp(value, tz=PKT).date()

    raw = clean(value)
    if not raw:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%d-%m-%Y",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue

    raise ValueError("Please select a valid date from the calendar.")


def resolve_click_clock_time(mode, date_value, hour_value, minute_value, ampm_value):
    """Calculates precise datetime object from mouse clock controls."""
    if mode != MANUAL_TIME_MODE:
        return get_current_time()

    selected_date = _picker_date_to_date(date_value)
    if selected_date is None:
        raise ValueError("Please select the date from the calendar.")

    hour_text = clean(hour_value)
    minute_text = clean(minute_value)
    ampm = clean(ampm_value).upper()

    if not hour_text or not minute_text or ampm not in AMPM_CHOICES:
        raise ValueError("Please select hour, minute and AM/PM using the clock controls.")

    hour12 = int(hour_text)
    minute = int(minute_text)
    if not (1 <= hour12 <= 12):
        raise ValueError("Hour must be between 1 and 12.")
    if not (0 <= minute <= 59):
        raise ValueError("Minute must be between 00 and 59.")

    hour24 = hour12 % 12
    if ampm == "PM":
        hour24 += 12

    return datetime(
        selected_date.year,
        selected_date.month,
        selected_date.day,
        hour24,
        minute,
        0,
        tzinfo=PKT,
    )


def clock_card_html(hour_value=None, minute_value=None, ampm_value=None, date_value=None):
    """Renders dynamic analog clock SVG preview card with digital time display."""
    now = get_current_time()

    try:
        hour12 = int(clean(hour_value)) if clean(hour_value) else ((now.hour - 1) % 12) + 1
    except ValueError:
        hour12 = ((now.hour - 1) % 12) + 1

    try:
        minute = int(clean(minute_value)) if clean(minute_value) else now.minute
    except ValueError:
        minute = now.minute

    ampm = clean(ampm_value).upper() if clean(ampm_value) else now.strftime("%p")
    if ampm not in AMPM_CHOICES:
        ampm = now.strftime("%p")

    minute_angle = minute * 6
    hour_angle = (hour12 % 12) * 30 + minute * 0.5

    def endpoint(cx, cy, length, angle_deg):
        angle = math.radians(angle_deg - 90)
        return cx + length * math.cos(angle), cy + length * math.sin(angle)

    cx, cy = 92, 72
    hx, hy = endpoint(cx, cy, 30, hour_angle)
    mx, my = endpoint(cx, cy, 43, minute_angle)

    nums = []
    for n in range(1, 13):
        x, y = endpoint(cx, cy, 55, n * 30)
        nums.append(
            f'<text x="{x:.1f}" y="{y + 4:.1f}" text-anchor="middle" '
            'font-size="11" font-weight="700" fill="#fff">'
            f'{n}</text>'
        )

    date_text = ""
    try:
        d = _picker_date_to_date(date_value)
        if d:
            date_text = d.strftime("%d-%m-%Y")
    except Exception:
        date_text = ""

    digital = f"{hour12:02d}:{minute:02d} {ampm}"
    subtitle = f"{date_text} · Pakistan (PKT)" if date_text else "Pakistan (PKT)"

    return f'''<div style="display:flex;align-items:center;gap:24px;background:#191919;color:white;border-radius:22px;padding:16px 24px;max-width:520px;min-height:150px;box-sizing:border-box;">
      <div style="min-width:190px;">
        <div style="font-size:30px;font-weight:750;line-height:1.1;">{digital}</div>
        <div style="font-size:14px;color:#c9c9c9;margin-top:10px;">{subtitle}</div>
        <div style="font-size:12px;color:#9fdaff;margin-top:8px;">Mouse clock preview</div>
      </div>
      <svg width="184" height="144" viewBox="0 0 184 144" role="img" aria-label="Selected analogue time">
        <circle cx="92" cy="72" r="67" fill="#202020" stroke="#3a3a3a" stroke-width="2"/>
        {''.join(nums)}
        <line x1="92" y1="72" x2="{hx:.1f}" y2="{hy:.1f}" stroke="#e8e8e8" stroke-width="4" stroke-linecap="round"/>
        <line x1="92" y1="72" x2="{mx:.1f}" y2="{my:.1f}" stroke="#e8e8e8" stroke-width="3" stroke-linecap="round"/>
        <circle cx="92" cy="72" r="4" fill="#ff6b35"/>
      </svg>
    </div>'''


def current_clock_picker_values():
    """Resets clock picker state variables to current PKT time."""
    now = get_current_time().replace(second=0, microsecond=0)
    return (
        MANUAL_TIME_MODE,
        now.strftime("%Y-%m-%d"),
        now.strftime("%I"),
        now.strftime("%M"),
        now.strftime("%p"),
        clock_card_html(
            now.strftime("%I"),
            now.strftime("%M"),
            now.strftime("%p"),
            now.strftime("%Y-%m-%d"),
        ),
    )


def update_clock_preview(date_value, hour_value, minute_value, ampm_value):
    """Wrapper to update clock preview on dropdown value changes."""
    return clock_card_html(hour_value, minute_value, ampm_value, date_value)


# ============================================================
# 5. OPENING EMAIL ENGINE
# ============================================================
OPENING_ISSUES = [
    "Link is down.",
    "Service degradation.",
    "Intermittent connectivity.",
    "High latency observed.",
    "Packet loss observed.",
    "Slow browsing issue.",
    "Bandwidth issue.",
    "Website / application issue.",
    "Other / Manual",
]


def generate_opening_email(
    selected_complaint,
    label,
    issue_type,
    custom_issue,
    ticket,
    priority,
    opening_time_mode,
    opening_date,
    opening_hour,
    opening_minute,
    opening_ampm,
):
    """Generates standardized Opening Acknowledgement email template."""
    label = resolve_complaint_label(selected_complaint, label)
    if not label:
        return "", "", "", "⚠️ Please select a complaint or enter the service label."

    issue = (
        clean(custom_issue) or "Service issue reported."
        if issue_type == "Other / Manual"
        else issue_type
    )

    try:
        requested_time = resolve_click_clock_time(
            opening_time_mode,
            opening_date,
            opening_hour,
            opening_minute,
            opening_ampm,
        )
    except (ValueError, TypeError) as exc:
        return label, "", "", f"⚠️ Invalid opening date/time: {exc}"

    service_type = detect_service_type(label)
    now = get_current_time()
    record = incident_records.get(label)
    reused_existing_time = False

    if record and record.get("status") == "OPEN" and record.get("reported_time"):
        reported_time = record["reported_time"]
        reused_existing_time = True
    else:
        reported_time = requested_time
        incident_records[label] = {
            "added_time": record.get("added_time") if record else now,
            "reported_time": reported_time,
            "restoration_time": None,
            "service_type": (
                record.get("service_type")
                if record and record.get("service_type")
                else service_type
            ),
            "issue_type": issue,
            "ticket": clean(ticket) or (record.get("ticket", "") if record else ""),
            "status": "OPEN",
            "last_stage": "Initial response",
        }

    record = incident_records[label]
    record["issue_type"] = issue
    record["ticket"] = clean(ticket) or record.get("ticket", "")
    record["last_stage"] = "Initial response"
    record["status"] = "OPEN"
    record["restoration_time"] = None

    subject = make_subject(label, issue.rstrip("."), priority, record.get("ticket", ""))

    email = f"""Dear Customer,

We acknowledge receipt of your complaint regarding service degradation/impact on your {record.get("service_type", service_type)} service. Our Corporate NOC has initiated an investigation.

Details:
Service Details: {label}
Issue Type: {issue}
Incident Start Time: {format_dt(reported_time)}
Current Status: Under investigation
Reference Ticket: {record.get("ticket", "")}

Our technical teams are actively working to identify the root cause and restore service at the earliest possible time. Regular updates will be shared until the issue is fully resolved.

We appreciate your patience and cooperation."""

    if reused_existing_time:
        runtime = (
            f"ℹ️ Complaint was already OPEN. Existing opening time retained: "
            f"{format_dt(reported_time)}."
        )
    else:
        source_note = (
            "manual mouse clock"
            if opening_time_mode == MANUAL_TIME_MODE
            else "current Pakistan time"
        )
        runtime = (
            f"✅ Complaint OPENED at {format_dt(reported_time)} "
            f"for {extract_service_id(label)} ({source_note})."
        )

    return label, subject, email, runtime


# ============================================================
# 6. CUSTOMER-END & FINDINGS ENGINE
# ============================================================
FINDING_OPTIONS = [
    "No problematic alarm observed at Node end",
    "Port is up",
    "Tunnel is intact",
    "ARP is not observed",
    "MAC is not observed",
    "User is offline at BRAS",
    "Optical power is optimal",
    "Traffic is observed at our end",
    "Latency is optimal",
    "No abnormality observed in transmission path",
]

CUSTOMER_ACTIONS = [
    "Verify last-mile media",
    "Check CPE/router and power",
    "Check SFP/optical module and equipment port",
    "Reconnect the affected PPPoE user",
    "Test through standalone laptop/device",
    "Bypass LAN and perform direct testing",
    "Share active POC for joint troubleshooting",
    "Share problematic stats for further investigation",
]


def finding_bullet(finding, service_type):
    if finding == "Traffic is observed at our end" and service_type == "Turbonet":
        return "Traffic is being observed at our end."
    return finding + "."


def customer_action_sentence(action):
    mapping = {
        "Verify last-mile media": "Kindly verify the service from the last-mile media and share your feedback.",
        "Check CPE/router and power": "Kindly verify the CPE/router status and power availability at your end.",
        "Check SFP/optical module and equipment port": "Kindly check the SFP/optical module and equipment port at your end and share updated status.",
        "Reconnect the affected PPPoE user": "Kindly reconnect the affected PPPoE user and confirm whether the session comes online.",
        "Test through standalone laptop/device": "Kindly perform testing through a standalone laptop/device to isolate any internal LAN impact.",
        "Bypass LAN and perform direct testing": "Kindly bypass the internal LAN and perform direct testing from the edge device.",
        "Share active POC for joint troubleshooting": "Kindly share an active POC so joint troubleshooting can be arranged.",
        "Share problematic stats for further investigation": "If the issue persists, kindly share the problematic statistics for further investigation.",
    }
    return mapping.get(action, "")


def generate_customer_end_email(
    label,
    issue_summary,
    findings,
    requested_action,
    priority,
    ticket,
):
    """Generates response email when issue is identified at customer side."""
    label = clean(label)
    if not label:
        return "", "⚠️ Please enter the service label.", ""

    service_type = detect_service_type(label)
    subject = make_subject(
        label,
        issue_summary or "Customer-End Verification",
        priority,
        ticket,
    )

    lines = [
        "Dear Customer,",
        "",
        "We have thoroughly checked the reported service.",
        "",
        "Findings are shared below:",
    ]

    if findings:
        for item in findings:
            lines.append(f"• {finding_bullet(item, service_type)}")
    else:
        lines.append("• No abnormality is currently observed at our end.")

    action_line = customer_action_sentence(requested_action)
    if action_line:
        lines.extend(["", action_line])

    lines.extend([
        "",
        "Please share your feedback so we can proceed accordingly.",
        "",
        "Your cooperation in this matter will be highly appreciated.",
    ])

    policy = ""
    if service_type == "Turbonet":
        policy = (
            "⚠️ Internal Turbo policy: Do not share the Turbonet graph with "
            "the customer. For graph-related requests, route the matter through "
            "KAM/GCSSQ."
        )

    return subject, "\n".join(lines), policy


# ============================================================
# 7. VENDOR & INTERNAL ESCALATION ENGINE (CC ROUTING & MANUAL OVERRIDE)
# ============================================================
ESCALATION_TARGETS = [
    "Netsat",
    "Comstar",
    "VT",
    "GCS",
    "WISS",
    "Nexlinx",
    "Connect",
    "Other",
]

ESCALATION_LEVELS = [
    "Initial engagement",
    "Follow-up",
    "Urgent",
    "Critical",
    "Multiple follow-ups / no response",
    "Request exact delay reason / RFO",
]

ESCALATION_FORMAT_OPTIONS = [
    "Option A (Standard Matrix Format)",
    "Option B (Subject-Linked Format)",
    "Option C ([Ticket] : [Subject] : [Vendor])",
]


def get_vendor_l1_l2_emails(vendor_name):
    """Extracts only Level 1 and Level 2 emails for a vendor (semicolon separated)."""
    if not vendor_name or vendor_name not in vendors_matrix_db:
        return ""
    emails = []
    for row in vendors_matrix_db[vendor_name]:
        lvl = str(row.get("Level", "")).upper()
        if "LEVEL 1" in lvl or "LEVEL 2" in lvl or "L1" in lvl or "L2" in lvl:
            raw_email = row.get("Email", "")
            for e in re.split(r"[;, ]+", raw_email):
                clean_e = e.strip()
                if clean_e and clean_e not in emails:
                    emails.append(clean_e)
    return "; ".join(emails)


def generate_outlook_eml_file(to_emails, cc_emails, subject, body_text, matrix_html):
    """Generates a standard .eml file with CC header that opens directly in Microsoft Outlook."""
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = "noc@company.com"
    if to_emails:
        msg["To"] = to_emails
    if cc_emails:
        msg["Cc"] = cc_emails

    msg.set_content(body_text)

    # Attach HTML version including the stylized escalation table
    html_body = f"""<html>
    <body style="font-family: Calibri, Arial, sans-serif; font-size: 14px; color: #000;">
        <div style="white-space: pre-wrap; margin-bottom: 20px;">{body_text}</div>
        <hr style="border: 0; border-top: 1px solid #ccc; margin: 20px 0;">
        {matrix_html}
    </body>
    </html>"""
    msg.add_alternative(html_body, subtype="html")

    temp_dir = tempfile.gettempdir()
    clean_sub = re.sub(r'[\/:*?"<>|]', "_", subject)[:40].strip()
    eml_filename = f"Escalation_{clean_sub}_{datetime.now().strftime('%H%M%S')}.eml"
    eml_path = os.path.join(temp_dir, eml_filename)

    with open(eml_path, "wb") as f:
        f.write(msg.as_bytes())

    return eml_path


def generate_escalation_email(
    label,
    target,
    escalation_level,
    escalation_format,
    rcbs_group,
    custom_to,
    custom_cc,
    custom_c_subject,
    ticket,
    findings,
    requested_action,
    priority,
    custom_target="",
):
    """Generates vendor / internal team escalation email across Option A, B, and C formats,

    and auto-creates a downloadable Outlook (.eml) message file with proper CC headers.
    """
    label_clean = clean(label)
    target_name = custom_target.strip() if target == "Other" and custom_target else target
    if not target_name:
        target_name = "Vendor"

    # 1. Resolve To & CC (User input takes priority; fallback to auto-fetch)
    to_emails = clean(custom_to) or get_vendor_l1_l2_emails(target_name)
    cc_emails = clean(custom_cc) or RCBS_GROUPS.get(rcbs_group, "")

    # 2. Determine Subject based on Option A, B, C
    if "Option A" in escalation_format:
        topic = f"{escalation_level} - {target_name}"
        subject = make_subject(label_clean, topic, priority, ticket)
        intro_sentence = "Please check the below-mentioned service, as the customer is currently facing a connectivity issue."
    elif "Option B" in escalation_format:
        subject = f"{label_clean} || {target_name}" if label_clean else f"Service Issue || {target_name}"
        intro_sentence = "Please check the link mentioned in subject, as the customer is currently facing a connectivity issue."
    else:
        # Option C: [Ticket Number] : [Subject] : [Vendor Name]
        sub_text = clean(custom_c_subject) or label_clean or "Service Outage"
        tkt_text = clean(ticket) or "No-Ticket"
        subject = f"[{tkt_text}] : [{sub_text}] : [{target_name}]"
        intro_sentence = "Please check the below-mentioned service, as the customer is currently facing a connectivity issue."

    # 3. Assemble Body
    greeting = f"Dear {target_name} Team,"
    extra_parts = []

    intro_map = {
        "Urgent": "Kindly prioritize the reported issue on urgent basis. The corporate customer is currently experiencing service impact/outage.",
        "Critical": "The reported issue is critically impacting our corporate customer. Immediate engagement and restoration on top priority are required.",
        "Multiple follow-ups / no response": f"Despite repeated follow-ups, we have not received a progressive update from the {target_name} team. The continued delay is impacting customer communication and is not acceptable.",
        "Request exact delay reason / RFO": "We have repeatedly requested a specific reason for the delay / clear RFO; however, only generic updates have been received. Kindly share the exact reason, defined action plan, and expected restoration timeline.",
    }
    if escalation_level in intro_map:
        extra_parts.append(intro_map[escalation_level])

    if findings:
        extra_parts.append("\nCurrent Findings:")
        if isinstance(findings, str):
            for line in findings.splitlines():
                if line.strip():
                    extra_parts.append(f"• {line.strip()}")
        elif isinstance(findings, (list, tuple)):
            for item in findings:
                if item:
                    extra_parts.append(f"• {item}.")

    req_act_clean = clean(requested_action)
    if req_act_clean:
        extra_parts.append(f"\nRequired Action: {req_act_clean}")

    if escalation_level in {
        "Urgent",
        "Critical",
        "Multiple follow-ups / no response",
        "Request exact delay reason / RFO",
    }:
        extra_parts.append("\nKindly share a clear progressive update/action plan at the earliest.")

    extra_content_str = "\n".join(extra_parts)
    extra_section = f"\n{extra_content_str}\n" if extra_parts else "\n"

    body = f"""{greeting}

{intro_sentence}

Service Details: {label_clean or "NA"}

Kindly investigate the issue, restore the service on priority, and share the ETTR at the earliest.{extra_section}
Your prompt support and cooperation will be highly appreciated."""

    # 4. Render HTML Matrix Table
    matrix_html = render_escalation_matrix_html(target_name)

    # 5. Generate Outlook .eml File
    eml_file_path = generate_outlook_eml_file(to_emails, cc_emails, subject, body, matrix_html)

    return subject, to_emails, cc_emails, body, eml_file_path, matrix_html


# ============================================================
# 8. CLOSURE & RFO ENGINE
# ============================================================
RCA_OPTIONS = [
    "Fiber Break (Dual/Triple)",
    "Fiber Break (Single)",
    "Fiber Break (Spur)",
    "Power Issue",
    "Power Issue at PTN / Node",
    "Port Down",
    "Port Problematic",
    "Customer related issue",
    "Customer own last mile",
    "Customer end power issue",
    "Hardware Faulty",
    "Configurations Issue",
    "Link Choking / Over-utilization",
    "Frequency Interference",
    "Optical Power Issue",
    "Vendor / Upstream Issue",
    "Routing Issue",
    "Submarine Cable Cut",
    "No Issue Observed",
    "Other / Manual",
]

ISSUE_FOUND_AT = [
    "Customer",
    "CONVEX",
    "NSS TEAM",
    "PRODUCT TEAM",
    "SELLER",
    "NOMC - Transmission Optical",
    "Vendor",
    "Region - Field Operations (FOPs)",
    "NOMC - IP Core",
    "Uplink ISP",
    "NOMC - Transmission Microwave",
    "TXN PM",
    "TXN CM",
    "RCA Awaited",
    "Other",
]

AUTO_ACTIONS = {
    "Fiber Break (Dual/Triple)": "Affected fiber section was restored/spliced and service connectivity was normalized.",
    "Fiber Break (Single)": "Affected fiber section was restored/spliced and service connectivity was normalized.",
    "Fiber Break (Spur)": "Spur fiber fault was restored and service connectivity was normalized.",
    "Power Issue": "Power was restored at the affected site/node and service became operational.",
    "Power Issue at PTN / Node": "Power was restored at the affected PTN/node and service became operational.",
    "Port Down": "Port connectivity was restored and the service became operational.",
    "Port Problematic": "The problematic port/connectivity was rectified and service was restored.",
    "Customer related issue": "Service became operational after customer-end verification/restoration.",
    "Customer own last mile": "The customer/last-mile issue was rectified and service became operational.",
    "Customer end power issue": "Customer-end power was restored and service became operational.",
    "Hardware Faulty": "The faulty hardware/equipment was rectified/replaced and service was restored.",
    "Configurations Issue": "Required configuration was rectified and service was restored.",
    "Link Choking / Over-utilization": "Bandwidth/utilization was normalized and the service was verified.",
    "Frequency Interference": "Wireless parameters/path were optimized and service stability was restored.",
    "Optical Power Issue": "Optical parameters were normalized and service connectivity was restored.",
    "Vendor / Upstream Issue": "The upstream/vendor issue was resolved and service was restored.",
    "Routing Issue": "Routing/path configuration was rectified and service was restored.",
    "Submarine Cable Cut": "Upstream capacity/path was restored or stabilized after the submarine cable incident.",
    "No Issue Observed": "No activity was performed at our end; the service was already found operational.",
}


def generate_closure_email(
    selected_complaint,
    label,
    issue_found_at,
    custom_issue_found_at,
    root_cause_option,
    custom_root_cause,
    corrective_action_option,
    custom_corrective_action,
    final_status,
    priority,
    closure_time_mode,
    closure_date,
    closure_hour,
    closure_minute,
    closure_ampm,
):
    """Generates closure RFO email and calculates total impact duration automatically."""
    label = resolve_complaint_label(selected_complaint, label)
    if not label:
        return "", "", "", "⚠️ Please select an OPEN complaint or enter the service label."

    record = incident_records.get(label)
    service_type = (
        record.get("service_type")
        if record and record.get("service_type")
        else detect_service_type(label)
    )

    root_cause = (
        clean(custom_root_cause) or "No issue observed"
        if root_cause_option == "Other / Manual"
        else root_cause_option
    )

    found_at = (
        clean(custom_issue_found_at)
        if issue_found_at == "Other"
        else issue_found_at
    )

    if corrective_action_option == "Auto from Root Cause":
        corrective_action = AUTO_ACTIONS.get(
            root_cause,
            "Required corrective action was completed and service was restored.",
        )
    elif corrective_action_option == "Other / Manual":
        corrective_action = (
            clean(custom_corrective_action)
            or "Required corrective action was completed."
        )
    else:
        corrective_action = corrective_action_option

    try:
        restoration_time = resolve_click_clock_time(
            closure_time_mode,
            closure_date,
            closure_hour,
            closure_minute,
            closure_ampm,
        )
    except (ValueError, TypeError) as exc:
        return label, "", "", f"⚠️ Invalid closing date/time: {exc}"

    no_issue_root_causes = [
        "No Issue Observed",
        "No Issue Found after testing",
    ]

    customer_end_indicators = [
        "Customer related issue",
        "Customer own last mile",
        "Customer end power issue",
        "Customer",
    ]

    is_no_issue_at_noc = False

    if root_cause in no_issue_root_causes:
        is_no_issue_at_noc = True

    if root_cause_option == "Other / Manual" and clean(custom_root_cause):
        custom_rc_lower = clean(custom_root_cause).lower()
        if any(phrase in custom_rc_lower for phrase in ["no issue", "no fault", "no abnormality", "no problem"]):
            is_no_issue_at_noc = True

    if found_at in customer_end_indicators:
        is_no_issue_at_noc = True

    no_action_phrases = [
        "No activity performed at our end",
        "no activity was performed",
        "already found operational",
        "already operating normally",
    ]
    if any(phrase in corrective_action.lower() for phrase in no_action_phrases):
        is_no_issue_at_noc = True

    if record and record.get("reported_time"):
        reported_time = record["reported_time"]

        if restoration_time < reported_time:
            return (
                label,
                "",
                "",
                f"⚠️ Closing time ({format_dt(restoration_time)}) cannot be earlier than "
                f"the saved opening time ({format_dt(reported_time)}).",
            )

        formatted_reported = format_dt(reported_time)

        if is_no_issue_at_noc:
            duration = "NA"
        else:
            duration = calculate_duration(reported_time, restoration_time)

        record["status"] = "CLOSED"
        record["last_stage"] = "Closed"
        record["restoration_time"] = restoration_time
        record["root_cause"] = root_cause
        record["issue_found_at"] = found_at
        record["corrective_action"] = corrective_action
    else:
        formatted_reported = "NA"
        duration = "NA"

        if record:
            record["status"] = "CLOSED"
            record["last_stage"] = "Closed without saved opening time"
            record["restoration_time"] = restoration_time

    subject = make_subject(label, "Service Restored / Closure", priority)

    if is_no_issue_at_noc:
        email = f"""Dear Customer,

We have thoroughly checked the reported service and no abnormality is currently observed at our end.

Incident Summary:
Service Details: {label}
Issue Found At: {found_at or "NA"}
Root Cause: {root_cause}
Corrective Action: {corrective_action}
Reported Time: {formatted_reported}
Verification Time: {format_dt(restoration_time)}
Service Impact Duration: NA

Note: As no issue was identified within our network, the service impact duration is not applicable.

The service is currently operating normally. Kindly verify and confirm if performance is satisfactory at your end.

We regret any inconvenience caused and appreciate your cooperation.

{final_status}"""
    else:
        email = f"""Dear Customer,

We are pleased to inform you that your {service_type} service is fully operational.

Incident Summary:
Service Details: {label}
Issue Found At: {found_at or "NA"}
Root Cause: {root_cause}
Corrective Action: {corrective_action}
Reported Time: {formatted_reported}
Restoration Time: {format_dt(restoration_time)}
Service Impact Duration: {duration}

The service is now operating normally. Kindly verify and confirm if performance is satisfactory at your end.

We regret any inconvenience caused and appreciate your cooperation throughout the restoration process.

{final_status}"""

    if record and record.get("reported_time"):
        if is_no_issue_at_noc:
            runtime = (
                f"✅ Complaint CLOSED at {format_dt(restoration_time)}. "
                f"Service Impact Duration: NA (No issue found at NOC end)."
            )
        else:
            runtime = (
                f"✅ Complaint CLOSED at {format_dt(restoration_time)}. "
                f"Total impact duration: {duration}."
            )
    elif record:
        runtime = (
            f"ℹ️ Complaint CLOSED at {format_dt(restoration_time)}, "
            "but no opening time was saved, so duration is NA."
        )
    else:
        runtime = (
            "ℹ️ Closure generated without a runtime complaint record. "
            "Reported Time and Duration are NA."
        )

    return label, subject, email, runtime


# ============================================================
# 9. STATS & TROUBLESHOOTING ENGINE
# ============================================================
STATS_SCENARIOS = [
    "Packet Loss / High Latency",
    "Slow Speed / Throughput",
    "Turbo Slow Speed",
    "Website Issue",
    "Particular Website / URL Issue",
    "Banking Application Issue",
    "TikTok Issue",
    "WhatsApp / Meta Issue",
    "PPPoE / BRAS Issue",
    "Voice / PRI / SIP Issue",
    "Optical Power Issue",
    "Routing / BGP Issue",
]


def stats_items(scenario):
    mapping = {
        "Packet Loss / High Latency": [
            "Ping statistics from the affected user",
            "WinMTR / traceroute towards the problematic destination",
            "Source IP / testing IP pool",
            "Destination IP / URL",
            "Affected user details",
            "Exact issue occurrence date and time",
        ],
        "Slow Speed / Throughput": [
            "Speedtest result",
            "Speedtest server details",
            "Affected user/source IP",
            "Direct/standalone testing result after bypassing the LAN",
            "Current issue occurrence time",
        ],
        "Turbo Slow Speed": [
            "Direct speed and browsing test through a standalone laptop/device",
            "Affected PPPoE user / source IP",
            "Speedtest server details and result",
            "Issue occurrence time",
        ],
        "Website Issue": [
            "Source IP pool",
            "Browser error screenshot",
            "Traceroute / WinMTR towards the problematic destination",
            "NSLOOKUP result of the problematic destination",
            "NSLOOKUP result of whoami.akamai.com",
            "Tracetcp towards the affected URL on port 443",
            "Primary and secondary DNS configuration",
        ],
        "Particular Website / URL Issue": [
            "Snapshot of whatismyipaddress.com",
            "Source IP pool",
            "Browser error screenshot",
            "Traceroute towards the problematic destination",
            "WinMTR / tracetcp towards the affected URL on port 443",
            "NSLOOKUP result of the problematic destination",
            "Primary and secondary DNS configuration",
        ],
        "Banking Application Issue": [
            "Snapshot of whatismyipaddress.com",
            "Source IP pool",
            "Browser/application error screenshot",
            "NSLOOKUP / traceroute / WinMTR details where applicable",
            "PCAP trace from the affected Android phone using PCAPdroid",
        ],
        "TikTok Issue": [
            "DNS server IP configured at the edge device",
            "NSLOOKUP whoami.akamai.com screenshot",
            "Standalone test using Ethernet cable and a test PPPoE user",
            "Windows Task Manager > Performance > Ethernet screenshot while playing a TikTok video",
            "Affected user/source IP and issue occurrence time",
        ],
        "WhatsApp / Meta Issue": [
            "Affected user/source IP",
            "Issue occurrence time",
            "Ping / WinMTR towards a reachable problematic destination if available",
            "DNS configuration",
            "Standalone testing result after bypassing the LAN",
        ],
        "PPPoE / BRAS Issue": [
            "PPPoE username",
            "Affected user MAC address",
            "CPE/router status",
            "Exact disconnect/login time",
            "Standalone dial/test result",
            "Screenshot/error message if authentication fails",
        ],
        "Voice / PRI / SIP Issue": [
            "Party-A number",
            "Party-B number",
            "Exact call date and time",
            "Incoming / outgoing call scenario",
            "Observed error / announcement",
            "CLI displayed at Party-B where applicable",
        ],
        "Optical Power Issue": [
            "Current TX/RX optical power readings",
            "SFP/optical module details",
            "Equipment port status",
            "Patch-cord/fiber inspection result",
            "Relevant alarm/screenshot for reference",
        ],
        "Routing / BGP Issue": [
            "Source IP pool / affected prefix",
            "Problematic destination IP / URL",
            "Traceroute / WinMTR",
            "Current route advertisement details",
            "Expected upstream/path if known",
        ],
    }
    return mapping.get(scenario, ["Relevant problematic statistics and timestamps"])


def generate_stats_request(label, scenario, audience, custom_context):
    """Generates required technical stats email request."""
    label = clean(label)
    subject = make_subject(
        label or scenario,
        f"Required Stats - {scenario}",
        "Normal",
    )

    lines = [
        salutation(audience),
        "",
        "To proceed with detailed investigation of the reported issue, kindly share the following information/statistics:",
        "",
    ]

    for item in stats_items(scenario):
        lines.append(f"• {item}")

    if clean(custom_context):
        lines.extend(["", clean(custom_context)])

    if scenario == "Banking Application Issue":
        lines.extend([
            "",
            "PCAPdroid capture method:",
            "Select PCAP file → select the target application → START capture → reproduce the banking-app issue/login error → STOP capture → share the generated PCAP file.",
        ])

    lines.extend([
        "",
        "Once the required information is received, we will proceed with further investigation accordingly.",
        "",
        "Your cooperation is highly appreciated.",
    ])

    policy = ""
    if scenario == "Turbo Slow Speed":
        policy = (
            "⚠️ Internal Turbo policy: Never share the Turbonet graph directly "
            "with the customer. If the customer requests the graph, route the "
            "request to KAM; graph-policy escalation should be handled with GCSSQ."
        )

    return subject, "\n".join(lines), policy


# ============================================================
# 10. PROGRESS UPDATE ENGINE
# ============================================================
PROGRESS_OPTIONS = [
    "Concerned team engaged",
    "Fault localization in progress",
    "Team dispatched",
    "Team reached site",
    "Fiber splicing in progress",
    "Rerouting in progress",
    "Configuration activity in progress",
    "Vendor engaged",
    "Customer coordination required",
    "Testing in progress",
    "Monitoring link stability",
]

PROGRESS_TEXT = {
    "Concerned team engaged": "Our concerned team is actively engaged and working on the reported issue.",
    "Fault localization in progress": "Fault localization is currently in progress and is being followed up on priority.",
    "Team dispatched": "The concerned field team has been dispatched for onsite troubleshooting/restoration.",
    "Team reached site": "The field team has reached the site and troubleshooting/restoration activity is in progress.",
    "Fiber splicing in progress": "Fiber splicing activity is currently in progress at the affected section.",
    "Rerouting in progress": "Rerouting activity is in progress to restore the affected service through an alternate path.",
    "Configuration activity in progress": "The concerned technical team is carrying out the required configuration activity.",
    "Vendor engaged": "The concerned vendor/upstream team has been engaged and is investigating the issue.",
    "Customer coordination required": "Further troubleshooting requires customer-side coordination. Kindly ensure an active POC is available.",
    "Testing in progress": "Detailed testing is currently in progress with the concerned teams.",
    "Monitoring link stability": "The link is currently under monitoring to verify stability and service performance.",
}


def generate_progress_email(
    label,
    progress_status,
    ettr_choice,
    custom_ettr,
    custom_note,
    priority,
    audience,
):
    """Generates progress status update email."""
    label = clean(label)
    if not label:
        return "", "⚠️ Please enter the service label."

    update_incident_stage(label, progress_status)
    subject = make_subject(label, f"Progress Update - {progress_status}", priority)

    body = f"""{salutation(audience)}

Please find the latest update regarding the reported service:

{PROGRESS_TEXT.get(progress_status, progress_status)}"""

    if clean(custom_note):
        body += f"\n\n{clean(custom_note)}"

    ettr_line = format_ettr(ettr_choice, custom_ettr)
    if ettr_line:
        body += f"\n\n{ettr_line}"

    body += (
        "\n\nFurther updates will be shared accordingly."
        "\n\nYour patience and cooperation are highly appreciated."
    )

    return subject, body


# ============================================================
# 11. OUTAGE ANALYZER ENGINE (STREAMLINED BGP MATCHING)
# ============================================================
outage_history = []


def extract_bw_numeric(label):
    match = re.search(r'(\d+(?:\.\d+)?)\s*(?:M|G|K)BPS', str(label).replace('Mpbs', 'Mbps'), re.IGNORECASE)
    if match:
        val = float(match.group(1))
        if 'G' in match.group(0).upper():
            val *= 1000
        return val
    return 0


def extract_bw_text(label):
    match = re.search(r'(\d+(?:\.\d+)?)\s*((?:M|G|K)BPS)', str(label).replace('Mpbs', 'Mbps'), re.IGNORECASE)
    if match:
        return f"{match.group(1)} {match.group(2).upper()}"
    return "NA"


def detect_category(label):
    """Streamlined category detection separating BGP DIA from DIA and transport."""
    lbl = str(label).upper()

    # 1. BGP DIA (Prioritized: matches any link containing BGP)
    if "BGP" in lbl:
        return "BGPDIA"

    # 2. Broadband / Access
    if "TURBO" in lbl or "TURBONET" in lbl:
        return "Turbonet"

    # 3. Transport & Dedicated Circuits
    if "MPLS" in lbl or "VPLS" in lbl:
        return "MPLS"
    if "DPLC" in lbl or "EPL" in lbl:
        return "DPLC"
    if "IPLC" in lbl:
        return "IPLC"
    if "M2M" in lbl:
        return "M2M"

    # 4. Standard DIA
    if "DIA" in lbl:
        return "DIA"

    # 5. Voice / Signaling
    if "SIP" in lbl or "PRI" in lbl:
        return "SIP/PRI"

    return "Other"


def extract_client_name(label):
    parts = [p.strip() for p in str(label).strip().split("_") if p.strip()]
    if parts and "ESSCLIENT" in parts[0].upper():
        parts = parts[1:]
    bw_idx = next((i for i, p in enumerate(parts) if re.search(r'\d+(?:\.\d+)?\s*(?:M|G|K)BPS', p.replace('Mpbs', 'Mbps'), re.I)), None)
    c_parts = parts[:bw_idx] if bw_idx is not None else (parts[:-1] if len(parts) > 1 else parts)
    control_words = {"DIA", "MPLS", "DPLC", "M2M", "TURBO", "TURBONET", "BGP", "BGPDIA", "DIABGP", "SIP", "PRI", "SIPPRI", "CENTRAL", "NORTH", "SOUTH"}
    clean_parts = [p for p in c_parts if p.upper() not in control_words and not p.isdigit()]
    return " ".join(clean_parts[:-1] if len(clean_parts) >= 2 else clean_parts).strip()


def extract_unique_id(label):
    parts = str(label).split('_')
    return parts[-1].strip() if len(parts) > 1 else ""


def load_and_clean_data(file_path):
    if file_path.lower().endswith(".csv"):
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
        header_idx = 0
        for i, line in enumerate(lines):
            if "Location Info" in line and "Last Occurred (ST)" in line:
                header_idx = i
                break
        clean_csv_data = "".join(lines[header_idx:])
        return pd.read_csv(io.StringIO(clean_csv_data), sep=",", engine="python")
    else:
        df = pd.read_excel(file_path, header=None)
        header_idx_list = df[df.apply(lambda r: r.astype(str).str.contains("Location Info", case=False).any(), axis=1)].index
        if not header_idx_list.empty:
            header_idx = header_idx_list[0]
            df.columns = df.iloc[header_idx]
            df = df.iloc[header_idx + 1:].reset_index(drop=True)
        return df


def process_outage_file(file_obj):
    """Processes uploaded CSV/Excel alarm dump files and generates BOTH full & simple excel reports."""
    if file_obj is None:
        return "⚠️ Please upload an Excel or CSV file.", None, None, get_history_table(), generate_handover_summary()

    file_path = file_obj.name
    try:
        df = load_and_clean_data(file_path)
        df.columns = [str(c).strip() for c in df.columns]

        if "Location Info" not in df.columns or "Last Occurred (ST)" not in df.columns:
            return "❌ Error: 'Location Info' or 'Last Occurred (ST)' column not found.", None, None, get_history_table(), generate_handover_summary()

        processed = []
        for _, row in df.iterrows():
            raw_label = str(row["Location Info"]).strip()
            if pd.isna(raw_label) or raw_label.upper() == "NAN" or "VISIBILITY" in raw_label.upper() or not raw_label.upper().startswith("ESS"):
                continue

            processed.append({
                "Service": raw_label,
                "Bandwidth_Text": extract_bw_text(raw_label),
                "Numeric_BW": extract_bw_numeric(raw_label),
                "Client": extract_client_name(raw_label),
                "Category": detect_category(raw_label),
                "UniqueID": extract_unique_id(raw_label),
                "Last Occurred (ST)": row["Last Occurred (ST)"],
            })

        if not processed:
            return "❌ No valid ESS service records found in file.", None, None, get_history_table(), generate_handover_summary()

        df_unsorted = pd.DataFrame(processed).drop_duplicates(subset=["Service"], keep="first").reset_index(drop=True)
        df_sorted = df_unsorted.sort_values(by="Numeric_BW", ascending=False).reset_index(drop=True)

        df_sorted["Time_Parsed"] = pd.to_datetime(df_sorted["Last Occurred (ST)"], errors="coerce")
        valid_times = df_sorted["Time_Parsed"].dropna()
        outlook_time = valid_times.iloc[0].strftime("%Y-%m-%d %I:%M:%S %p") if not valid_times.empty else "Not Found"

        counts = Counter(df_sorted["Category"])
        summary = ", ".join([f"{c}x {s}" for s, c in sorted(counts.items(), key=lambda x: x[1], reverse=True)])

        priority_df = df_sorted[df_sorted["Numeric_BW"] > 250]
        normal_df = df_sorted[df_sorted["Numeric_BW"] <= 250]

        # Generate Console Text Output
        text_lines = []
        text_lines.append("=" * 80)
        text_lines.append(f"{'OUTLOOK READY TEXT':^80}")
        text_lines.append("=" * 80)
        text_lines.append(f"PTN Layer: {len(df_sorted)}x ({summary}) are down")
        text_lines.append("last occured time:")
        text_lines.append(outlook_time)
        text_lines.append("=" * 80 + "\n")

        text_lines.append("Transmission Section\n")
        text_lines.append("Priority Clients")

        for _, row in priority_df.iterrows():
            text_lines.append(row['Service'])

        text_lines.append("\n")

        for _, row in normal_df.iterrows():
            text_lines.append(row['Service'])

        text_lines.append("\nOutage Thread")
        text_lines.append("-" * 50)
        text_lines.append("\n")

        for _, row in df_unsorted.iterrows():
            text_lines.append(row['Service'])

        console_output = "\n".join(text_lines)

        # Append to Shift History Log
        now_str = datetime.now(PKT).strftime("%I:%M %p")
        outage_history.append({
            "id": len(outage_history) + 1,
            "processed_time": now_str,
            "file_name": os.path.basename(file_path),
            "total_links": len(df_sorted),
            "summary": summary,
            "priority_count": len(priority_df),
            "occurred_time": outlook_time,
        })

        file_time = valid_times.iloc[0].strftime("%Y-%m-%d_%I-%M_%p") if not valid_times.empty else datetime.now().strftime("%Y-%m-%d_%I-%M_%p")
        temp_dir = tempfile.gettempdir()

        # 1. Full Categorized Report
        out_filepath = os.path.join(temp_dir, f"Outage_Links_{file_time}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Categorized Links"

        fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        font = Font(bold=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        ws.append(["Part 1: Raw Unsorted Data"])
        ws.cell(row=1, column=1).font = Font(bold=True)
        headers = ["Service", "Bandwidth", "Client", "Category", "UniqueID", "Last Occurred (ST)"]
        ws.append(headers)

        for c in ws[2]:
            c.fill = fill
            c.font = font
            c.border = border

        for _, r in df_unsorted.iterrows():
            ws.append([r["Service"], r["Bandwidth_Text"], r["Client"], r["Category"], r["UniqueID"], r["Last Occurred (ST)"]])

        gap_row = ws.max_row + 3
        ws.cell(row=gap_row, column=1, value="Part 2: Sorted Data for Transmission Rerouting (Descending Bandwidth)")
        ws.cell(row=gap_row, column=1).font = Font(bold=True, italic=True)

        header_row = gap_row + 1
        ws.append(headers)
        for c in ws[header_row]:
            c.fill = fill
            c.font = font
            c.border = border

        for _, r in df_sorted.iterrows():
            bw_val = float(r["Numeric_BW"])
            bw_num = int(bw_val) if bw_val.is_integer() else bw_val
            if bw_num == 0: bw_num = "NA"
            time_fmt = r['Time_Parsed'].strftime("%Y-%m-%d %I:%M:%S %p") if pd.notna(r['Time_Parsed']) else r["Last Occurred (ST)"]
            ws.append([r["Service"], bw_num, r["Client"], r["Category"], r["UniqueID"], time_fmt])

        for col_idx, width in enumerate([75, 15, 35, 15, 25, 25], start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        wb.save(out_filepath)

        # 2. Simple Outage Links Report
        simple_filepath = os.path.join(temp_dir, f"Sample Links.xlsx")
        wb_simple = Workbook()
        ws_simple = wb_simple.active
        ws_simple.title = "Outage Links"

        ws_simple.append(["Links"])
        ws_simple.cell(row=1, column=1).font = Font(bold=True)

        for _, r in df_unsorted.iterrows():
            ws_simple.append([r["Service"]])

        ws_simple.column_dimensions['A'].width = 80
        wb_simple.save(simple_filepath)

        return console_output, out_filepath, simple_filepath, get_history_table(), generate_handover_summary()

    except Exception as e:
        return f"❌ Processing Error: {str(e)}", None, None, get_history_table(), generate_handover_summary()


def get_history_table():
    rows = []
    for item in reversed(outage_history):
        rows.append([
            f"Outage #{item['id']}",
            item['processed_time'],
            item['total_links'],
            item['priority_count'],
            item['summary'],
            item['occurred_time'],
        ])
    return rows


def generate_handover_summary():
    if not outage_history:
        return "No outages processed in the current shift history."

    total_outages = len(outage_history)
    total_links = sum(item['total_links'] for item in outage_history)
    total_priority = sum(item['priority_count'] for item in outage_history)

    lines = [
        "==================================================",
        "📋 SHIFT HANDOVER SUMMARY / DAILY OUTAGE REPORT",
        "==================================================",
        f"• Total Outages Handled Today: {total_outages}",
        f"• Total Down Links Processed: {total_links}",
        f"• High Priority Links (>250M): {total_priority}",
        "--------------------------------------------------",
        "DETAILS BREAKDOWN:",
    ]

    for item in outage_history:
        lines.append(
            f"• Outage #{item['id']} ({item['processed_time']}): "
            f"{item['total_links']} Links Down ({item['summary']}) | "
            f"Occurred: {item['occurred_time']}"
        )

    lines.append("==================================================")
    return "\n".join(lines)


def reset_outage_history():
    global outage_history
    outage_history = []
    return [], "No outages processed in the current shift history."


# ============================================================
# 12. VENDOR ESCALATION MATRIX MANAGER (JSON PERSISTENCE)
# ============================================================
VENDOR_JSON_FILE = "vendors.json"

DEFAULT_VENDOR_MATRIX = {
    "Netsat": [
        {"Level": "Level 1", "Name": "Netsat Support Desk", "Designation": "Help desk / CNOC Netsat", "Time": "1 Hour", "Phone": "021-111638728", "Email": "support.cmpak@netsat.net.pk"},
        {"Level": "Level 1", "Name": "Ali Nadeem", "Designation": "Help desk / CNOC Netsat", "Time": "1 Hour", "Phone": "0309 1112422; 0312-2431968", "Email": "ali.nadeem@netsat.net.pk"},
        {"Level": "Level 2 (Central Region)", "Name": "Zafar Iqbal (CTR)", "Designation": "Regional head (Lahore)", "Time": "3 hours", "Phone": "0301-8114185", "Email": "zafar.iqbal@netsat.net.pk"},
        {"Level": "Level 2 (Central Region)", "Name": "Shoukat Khan (MTR)", "Designation": "Regional head (Multan)", "Time": "3 hours", "Phone": "0302-8271762", "Email": "shoukat.khan@netsat.net.pk"},
        {"Level": "Level 2 (South Region)", "Name": "Farhan (Sindh)", "Designation": "Regional head (Sindh)", "Time": "3 hours", "Phone": "0301-8114182", "Email": "m.farhan@netsat.net.pk"},
        {"Level": "Level 2 (South Region)", "Name": "Syed Yousuf Hussain (Balochistan)", "Designation": "Regional head (Balochistan)", "Time": "3 hours", "Phone": "0300-8259632", "Email": "syed.yousuf@netsat.net.pk"},
        {"Level": "Level 2 (North Region)", "Name": "Asad Imran (North/KPK)", "Designation": "Regional head (North)", "Time": "3 hours", "Phone": "0311-8814294", "Email": "asad.imran@netsat.net.pk"},
        {"Level": "Level 3", "Name": "M. Tahir Qureshi", "Designation": "Operational Support", "Time": "6 hours", "Phone": "0300 8232647", "Email": "m.tahir@netsat.net.pk"},
        {"Level": "Level 3", "Name": "Muhammad Sumair", "Designation": "Operation Head", "Time": "8 hours", "Phone": "0301-8114181", "Email": "sumair@netsat.net.pk"},
        {"Level": "Level 4", "Name": "Syed Mubashir Imam", "Designation": "Senior Operational Head", "Time": "12 hours", "Phone": "0301-8292632", "Email": "mubashir.imam@netsat.net.pk"},
        {"Level": "Level 5", "Name": "Kamran Qaiser", "Designation": "HOD Wireless", "Time": "Urgent/Emergency Maintenance", "Phone": "0301 8114177", "Email": "kamran.qaiser@netsat.net.pk"},
    ],
    "Comstar": [
        {"Level": "Level 1", "Name": "Support Desk", "Designation": "CS", "Time": "10-15min", "Phone": "0333-1312343", "Email": "cs@comstar.com.pk"},
        {"Level": "Level 2", "Name": "Abdul Wajid", "Designation": "TL (South)", "Time": "15-30min", "Phone": "0334-2594529", "Email": "awajid@comstar.com.pk"},
        {"Level": "Level 3", "Name": "Osman Javaid", "Time": "30-45min", "Designation": "Manager Engineering", "Phone": "0336-5479293", "Email": "ojavaid@comstar.com.pk"},
    ],
    "Vision Telecom": [
        {"Level": "Level 1", "Name": "Corporate Helpdesk", "Designation": "Support", "Time": "Immediate", "Phone": "0308-8881418", "Email": "support@visiontelecom.com.pk"},
        {"Level": "Level 2", "Name": "Rizwan Younis", "Designation": "Team Lead", "Time": "30min", "Phone": "0300-0807140", "Email": "rizwan.younis@visiontelecom.com.pk"},
    ],
}


def load_vendors_matrix():
    """Loads vendor matrix dictionary directly from JSON file, initializing defaults if missing."""
    if not os.path.exists(VENDOR_JSON_FILE):
        save_vendors_matrix(DEFAULT_VENDOR_MATRIX)
        return DEFAULT_VENDOR_MATRIX
    try:
        with open(VENDOR_JSON_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if data else DEFAULT_VENDOR_MATRIX
    except Exception:
        return DEFAULT_VENDOR_MATRIX


def save_vendors_matrix(data):
    """Saves updated vendor matrix dictionary directly to vendors.json."""
    try:
        with open(VENDOR_JSON_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving vendors.json: {e}")


vendors_matrix_db = load_vendors_matrix()


def get_vendor_list():
    """Returns dynamic list of all registered vendor names."""
    return list(vendors_matrix_db.keys())


def get_vendor_emails_string(vendor_name):
    """Extracts all emails for a vendor into semicolon-separated string for Outlook."""
    if not vendor_name or vendor_name not in vendors_matrix_db:
        return ""
    emails = []
    for row in vendors_matrix_db[vendor_name]:
        raw_email = row.get("Email", "")
        for e in re.split(r"[;, ]+", raw_email):
            clean_e = e.strip()
            if clean_e and clean_e not in emails:
                emails.append(clean_e)
    return "; ".join(emails)


def render_escalation_matrix_html(vendor_name):
    """Renders HTML matrix table with corporate styling and email CSS selection."""
    if not vendor_name or vendor_name not in vendors_matrix_db:
        return "<p style='color: gray;'>No vendor selected or matrix empty.</p>"

    matrix = vendors_matrix_db[vendor_name]

    html = f"""
    <style>
        .matrix-table,
        .matrix-table th,
        .matrix-table td {{
            -webkit-user-select: none !important;
            -moz-user-select: none !important;
            -ms-user-select: none !important;
            user-select: none !important;
        }}

        .matrix-table td.email-cell,
        .matrix-table td.email-cell a {{
            -webkit-user-select: text !important;
            -moz-user-select: text !important;
            -ms-user-select: text !important;
            user-select: text !important;
            cursor: text !important;
            color: #0000ee !important;
            font-weight: bold;
        }}
    </style>

    <div style="font-family: Arial, sans-serif; max-width: 100%; overflow-x: auto; margin-top: 10px;">
        <table class="matrix-table" style="width: 100%; border-collapse: collapse; border: 1.5px solid #000; font-size: 13px; text-align: left; background:#fff;">
            <thead>
                <tr style="background-color: #7b241c; color: white;">
                    <th colspan="6" style="padding: 10px; font-size: 15px; font-weight: bold; border-bottom: 2px solid #000; text-align:center;">
                        Escalation Matrix - {vendor_name}
                    </th>
                </tr>
                <tr style="background-color: #922b21; color: white; border-bottom: 1.5px solid #000;">
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 18%;">Escalation levels</th>
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 20%;">Name</th>
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 20%;">Designation</th>
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 14%;">Escalation Time</th>
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 14%;">Contact Number</th>
                    <th style="padding: 8px; border: 1px solid #7b241c; width: 14%;">Email address</th>
                </tr>
            </thead>
            <tbody>
    """

    for row in matrix:
        email_val = row.get('Email', '')
        email_html = f'<a href="mailto:{email_val}">{email_val}</a>' if email_val else ''
        html += f"""
        <tr style="border: 1px solid #ccc; background-color: #ffffff; color:#000;">
            <td style="padding: 7px; border: 1px solid #333; font-weight: bold;">{row.get('Level', '')}</td>
            <td style="padding: 7px; border: 1px solid #333;">{row.get('Name', '')}</td>
            <td style="padding: 7px; border: 1px solid #333;">{row.get('Designation', '')}</td>
            <td style="padding: 7px; border: 1px solid #333;">{row.get('Time', '')}</td>
            <td style="padding: 7px; border: 1px solid #333;">{row.get('Phone', '')}</td>
            <td class="email-cell" style="padding: 7px; border: 1px solid #333;">{email_html}</td>
        </tr>
        """

    html += """
            </tbody>
        </table>
    </div>
    """
    return html


def get_matrix_dataframe(vendor_name):
    if not vendor_name or vendor_name not in vendors_matrix_db:
        return []
    return [[r.get("Level", ""), r.get("Name", ""), r.get("Designation", ""), r.get("Time", ""), r.get("Phone", ""), r.get("Email", "")] for r in vendors_matrix_db[vendor_name]]


def add_vendor_name(new_vendor_name):
    """Adds a new vendor entry and syncs to vendors.json."""
    if not new_vendor_name or not new_vendor_name.strip():
        return gr.update(choices=get_vendor_list()), "", "❌ Vendor name is required!"
    v_name = new_vendor_name.strip()
    if v_name not in vendors_matrix_db:
        vendors_matrix_db[v_name] = []
        save_vendors_matrix(vendors_matrix_db)
    return gr.update(choices=get_vendor_list(), value=v_name), get_vendor_emails_string(v_name), f"✅ Vendor '{v_name}' added & saved to vendors.json!"


def add_contact_to_matrix(vendor_name, level, name, designation, time, phone, email):
    """Adds contact person to specified escalation level and auto-saves to vendors.json."""
    if not vendor_name:
        return render_escalation_matrix_html(vendor_name), get_matrix_dataframe(vendor_name), get_vendor_emails_string(vendor_name), "⚠️ Please select a Vendor first!"

    if not name or not email:
        return render_escalation_matrix_html(vendor_name), get_matrix_dataframe(vendor_name), get_vendor_emails_string(vendor_name), "❌ Name and Email are required!"

    vendors_matrix_db[vendor_name].append({
        "Level": level,
        "Name": name.strip(),
        "Designation": designation.strip(),
        "Time": time.strip(),
        "Phone": phone.strip(),
        "Email": email.strip(),
    })

    save_vendors_matrix(vendors_matrix_db)

    return (
        render_escalation_matrix_html(vendor_name),
        get_matrix_dataframe(vendor_name),
        get_vendor_emails_string(vendor_name),
        f"✅ Contact added & saved to vendors.json for {vendor_name}!",
    )


def delete_contact_from_matrix(vendor_name, name_to_delete):
    """Deletes contact person from matrix and auto-saves to vendors.json."""
    if not vendor_name or vendor_name not in vendors_matrix_db:
        return render_escalation_matrix_html(vendor_name), get_matrix_dataframe(vendor_name), get_vendor_emails_string(vendor_name), "⚠️ Vendor not found!"

    vendors_matrix_db[vendor_name] = [r for r in vendors_matrix_db[vendor_name] if r.get("Name") != name_to_delete.strip()]
    save_vendors_matrix(vendors_matrix_db)

    return (
        render_escalation_matrix_html(vendor_name),
        get_matrix_dataframe(vendor_name),
        get_vendor_emails_string(vendor_name),
        f"🗑️ Deleted '{name_to_delete}' & saved changes to vendors.json!",
    )


# ============================================================
# 13. ACTIVE DASHBOARD & MONITORING
# ============================================================
def refresh_dashboard(view_mode):
    records = []

    for label, record in incident_records.items():
        status = record.get("status", "QUEUED")

        if view_mode == "Open Only" and status != "OPEN":
            continue
        if view_mode == "Queued Only" and status != "QUEUED":
            continue
        if view_mode == "Closed Only" and status != "CLOSED":
            continue
        if view_mode == "Queued + Open" and status not in ("QUEUED", "OPEN"):
            continue

        added_time = record.get("added_time")
        reported_time = record.get("reported_time")
        restoration_time = record.get("restoration_time")

        if reported_time:
            age = incident_age_text(
                reported_time,
                added_time=added_time,
                status=status,
            )
        else:
            age = incident_age_text(
                None,
                added_time=added_time,
                status=status,
            )

        duration = (
            calculate_duration(reported_time, restoration_time)
            if reported_time and restoration_time
            else ""
        )

        records.append([
            extract_service_id(label),
            record.get("service_type", ""),
            record.get("issue_type", ""),
            format_dt(added_time) if added_time else "",
            format_dt(reported_time) if reported_time else "Not opened",
            age,
            format_dt(restoration_time) if restoration_time else "",
            duration,
            record.get("ticket", ""),
            record.get("last_stage", ""),
            status,
        ])

    records.sort(
        key=lambda row: row[3] or "",
        reverse=True,
    )
    return records


def check_incident(label):
    label = clean(label)
    if not label:
        return "⚠️ Please enter/select a service label."

    if label not in incident_records:
        return "❌ No runtime complaint record found for this label."

    record = incident_records[label]
    added = record.get("added_time")
    reported = record.get("reported_time")
    restored = record.get("restoration_time")

    duration = (
        calculate_duration(reported, restored)
        if reported and restored
        else "Running" if reported else "Not started"
    )

    return f"""Status: {record.get("status", "QUEUED")}
Service Type: {record.get("service_type", "")}
Issue Type: {record.get("issue_type", "")}
Added to Queue: {format_dt(added) if added else "NA"}
Opening Time: {format_dt(reported) if reported else "Not opened"}
Current Age: {incident_age_text(reported, added, record.get("status", "QUEUED"))}
Closing Time: {format_dt(restored) if restored else "Not closed"}
Total Duration: {duration}
Last Stage: {record.get("last_stage", "")}
Reference Ticket: {record.get("ticket", "")}"""


def reset_incident(label):
    label = clean(label)
    if not label:
        return "⚠️ Please enter/select a service label."

    if label in incident_records:
        del incident_records[label]
        return "✅ Runtime complaint record cleared."

    return "❌ No runtime complaint found for this label."


# ============================================================
# 14. GRADIO UI BUILDER
# ============================================================
def build_app():
    with gr.Blocks(title="Corporate NOC Outage Reporting Console") as app:
        gr.Markdown(
            """
# 🌐 Corporate NOC Outage Reporting Console
Runtime-only console for parallel complaint handling, vendor escalations, troubleshooting, and daily handover reports.
            """
        )

        with gr.Row():
            label = gr.Textbox(
                label="Service Label / Link Name",
                placeholder="ESSClient_DIA_... / SITE-... / link name",
                lines=1,
            )
            common_ticket = gr.Textbox(
                label="Reference Ticket",
                placeholder="Optional",
                lines=1,
            )

        with gr.Tabs():
            # TAB 1: COMPLAINT MANAGER
            with gr.Tab("🗂️ Complaint Manager"):
                gr.Markdown("### Parallel Complaint Queue")

                bulk_complaints = gr.Textbox(
                    label="Add Multiple Complaints",
                    placeholder=(
                        "ESSClient_DPLC_A.A Network_Tandlianwala-FSD_900Mbps_DPLC14722SL1\n"
                        "ESSClient_Central_Turbo_AAA_Brouadband_Faisalabad_2_981Mbps_Turbo14648SL694\n"
                        "ESSClient_SIP_PRI_2_ACE Money Transfer_Kharian_2Mbps_SIP00042"
                    ),
                    lines=8,
                )

                with gr.Row():
                    complaint_default_issue = gr.Dropdown(
                        choices=COMPLAINT_DEFAULT_ISSUES,
                        value="Link is down.",
                        label="Default Complaint Type",
                    )
                    complaint_bulk_ticket = gr.Textbox(
                        label="Reference Ticket (Optional)",
                        placeholder="Applied to newly added complaints",
                    )

                add_complaints_button = gr.Button(
                    "Add Complaints to Runtime Queue",
                    variant="primary",
                )

                with gr.Row():
                    complaint_selector = gr.Dropdown(
                        choices=[],
                        label="Select Runtime Complaint",
                        info="Select a complaint, then load it into the main Service Label field.",
                    )
                    refresh_complaints_button = gr.Button("Refresh Complaint List")

                with gr.Row():
                    load_complaint_button = gr.Button("Load Selected Complaint", variant="primary")
                    remove_complaint_button = gr.Button("Remove Selected Complaint", variant="stop")

                complaint_manager_status = gr.Textbox(
                    label="Complaint Manager Status",
                    lines=5,
                    interactive=False,
                )

                complaint_manager_dashboard = gr.Dataframe(
                    headers=[
                        "Service", "Type", "Issue", "Added", "Opened",
                        "Age", "Closed", "Total Time", "Ticket", "Last Stage", "Status"
                    ],
                    datatype=["str"] * 11,
                    value=[],
                    interactive=False,
                    label="Queued + Open Complaints",
                )

            # TAB 2: OUTAGE ANALYZER
            with gr.Tab("📈 Outage Analyzer"):
                gr.Markdown("### 📊 NOC Alarm & Outage Link Analyzer")
                with gr.Row():
                    with gr.Column():
                        file_input = gr.File(
                            label="Upload Alarm Dump File (CSV / Excel)",
                            file_types=[".csv", ".xlsx", ".xls"],
                        )
                        analyze_button = gr.Button("⚡ Analyze Outage File", variant="primary")
                    with gr.Column():
                        output_file = gr.File(label="📥 Download Categorized Excel Report")
                        simple_links_file = gr.File(label="📄 Download Simple Outage Links Only")

                console_output = gr.Textbox(
                    label="📋 Outlook Ready Summary & Thread Output",
                    lines=14,
                    interactive=False,
                )

                gr.Markdown("---")
                gr.Markdown("### 📜 Shift Outage History & Handover Summary")

                with gr.Row():
                    clear_history_btn = gr.Button("🗑️ Clear Shift History", variant="stop")

                history_table = gr.Dataframe(
                    headers=[
                        "Outage ID", "Processed Time", "Total Links Down",
                        "Priority (>250M)", "Categorized Breakdown", "Last Occurred Time"
                    ],
                    datatype=["str", "str", "number", "number", "str", "str"],
                    value=[],
                    interactive=False,
                    label="Shift Outage Log",
                )

                handover_text = gr.Textbox(
                    label="🗣️ Shift Handover Communication Text (Copy for Next Technical Shift)",
                    lines=8,
                    interactive=False,
                )

                analyze_button.click(
                    fn=process_outage_file,
                    inputs=file_input,
                    outputs=[console_output, output_file, simple_links_file, history_table, handover_text],
                )

                clear_history_btn.click(
                    fn=reset_outage_history,
                    inputs=[],
                    outputs=[history_table, handover_text],
                )

            # TAB 3: VENDOR MATRIX MANAGER
            with gr.Tab("🏪 Vendor Matrix Manager"):
                gr.Markdown("### 📊 Vendor Escalation Matrix Management (JSON Persistence)")

                with gr.Row():
                    with gr.Column(scale=1):
                        gr.Markdown("#### 1️⃣ Select / Add Vendor")
                        vendor_select_dd = gr.Dropdown(
                            choices=get_vendor_list(),
                            value=get_vendor_list()[0] if get_vendor_list() else None,
                            label="Select Vendor",
                        )
                        new_vendor_tb = gr.Textbox(label="Or Add New Vendor Name", placeholder="e.g. Wateen / Cybernet")
                        add_vendor_btn = gr.Button("➕ Add Vendor Name")

                    with gr.Column(scale=2):
                        gr.Markdown("#### 2️⃣ Add Contact Level to Matrix")
                        with gr.Row():
                            level_dd = gr.Dropdown(
                                choices=[
                                    "Level 1",
                                    "Level 2 (Central Region)",
                                    "Level 2 (South Region)",
                                    "Level 2 (North Region)",
                                    "Level 3",
                                    "Level 4",
                                    "Level 5",
                                ],
                                value="Level 1",
                                label="Escalation Level",
                            )
                            c_name_tb = gr.Textbox(label="Contact Name", placeholder="e.g. Ali Nadeem")
                            c_desig_tb = gr.Textbox(label="Designation", placeholder="e.g. Regional Head")

                        with gr.Row():
                            c_time_tb = gr.Textbox(label="Escalation Time", placeholder="e.g. 1 Hour / 3 hours")
                            c_phone_tb = gr.Textbox(label="Contact Phone", placeholder="0300-XXXXXXX")
                            c_email_tb = gr.Textbox(label="Email Address", placeholder="person@vendor.com")

                        add_contact_btn = gr.Button("➕ Add Contact to Matrix", variant="primary")

                matrix_status_msg = gr.Textbox(label="Action Status", interactive=False)

                gr.Markdown("---")
                gr.Markdown("#### 📋 Copy All Emails for Outlook")
                copy_emails_tb = gr.Textbox(
                    label="All Vendor Emails (Copy-paste directly to Outlook To/CC)",
                    value=get_vendor_emails_string(get_vendor_list()[0] if get_vendor_list() else ""),
                    interactive=True,
                )

                gr.Markdown("---")
                gr.Markdown("### 📄 Live Matrix Table Preview (HTML Red Header Format)")

                initial_vendor = get_vendor_list()[0] if get_vendor_list() else ""
                matrix_html_preview = gr.HTML(value=render_escalation_matrix_html(initial_vendor))

                gr.Markdown("---")
                with gr.Row():
                    with gr.Column():
                        gr.Markdown("#### 🗑️ Delete Contact Person")
                        del_name_tb = gr.Textbox(label="Enter Exact Name to Delete", placeholder="e.g. Ali Nadeem")
                        del_contact_btn = gr.Button("🗑️ Remove Person", variant="stop")

                    with gr.Column():
                        gr.Markdown("#### 📋 Dataframe View")
                        matrix_df_view = gr.Dataframe(
                            headers=["Level", "Name", "Designation", "Time", "Phone", "Email"],
                            value=get_matrix_dataframe(initial_vendor),
                            interactive=False,
                        )

                vendor_select_dd.change(
                    fn=lambda v: (render_escalation_matrix_html(v), get_matrix_dataframe(v), get_vendor_emails_string(v)),
                    inputs=vendor_select_dd,
                    outputs=[matrix_html_preview, matrix_df_view, copy_emails_tb],
                )

                add_vendor_btn.click(
                    fn=add_vendor_name,
                    inputs=new_vendor_tb,
                    outputs=[vendor_select_dd, copy_emails_tb, matrix_status_msg],
                )

                add_contact_btn.click(
                    fn=add_contact_to_matrix,
                    inputs=[vendor_select_dd, level_dd, c_name_tb, c_desig_tb, c_time_tb, c_phone_tb, c_email_tb],
                    outputs=[matrix_html_preview, matrix_df_view, copy_emails_tb, matrix_status_msg],
                )

                del_contact_btn.click(
                    fn=delete_contact_from_matrix,
                    inputs=[vendor_select_dd, del_name_tb],
                    outputs=[matrix_html_preview, matrix_df_view, copy_emails_tb, matrix_status_msg],
                )

            # TAB 4: OPENING
            with gr.Tab("🚨 Opening"):
                gr.Markdown("### Open / Start a Complaint")

                with gr.Row():
                    opening_complaint_selector = gr.Dropdown(
                        choices=[],
                        label="Select Complaint to Open",
                    )
                    opening_refresh_complaints = gr.Button("Refresh Opening Complaint List")

                opening_selected_details = gr.Textbox(
                    label="Selected Complaint Details",
                    lines=3,
                    interactive=False,
                )

                opening_issue = gr.Dropdown(
                    choices=OPENING_ISSUES,
                    value="Link is down.",
                    label="Issue Type",
                )
                opening_custom_issue = gr.Textbox(
                    label="Custom Issue",
                    placeholder="Used only when Other / Manual is selected",
                )
                opening_priority = gr.Dropdown(
                    choices=["Normal", "Urgent", "Critical"],
                    value="Normal",
                    label="Priority",
                )

                gr.Markdown("### 🕐 Opening Clock")
                opening_time_mode = gr.Radio(
                    choices=[AUTO_TIME_MODE, MANUAL_TIME_MODE],
                    value=AUTO_TIME_MODE,
                    label="Opening Time Mode",
                )

                with gr.Row():
                    opening_date = gr.DateTime(
                        value=None,
                        include_time=False,
                        type="string",
                        label="Opening Date",
                        interactive=True,
                    )
                    opening_hour = gr.Dropdown(
                        choices=HOUR_CHOICES,
                        value=None,
                        label="Hour",
                    )
                    opening_minute = gr.Dropdown(
                        choices=MINUTE_CHOICES,
                        value=None,
                        label="Minute",
                    )
                    opening_ampm = gr.Radio(
                        choices=AMPM_CHOICES,
                        value=None,
                        label="AM / PM",
                    )

                opening_set_current = gr.Button("🕐 Set Manual Picker to Current Pakistan Time")
                opening_clock_preview = gr.HTML(
                    value=clock_card_html(),
                    label="Opening Clock Preview",
                )

                opening_button = gr.Button("Generate Opening Email", variant="primary")
                opening_subject = gr.Textbox(label="Subject", interactive=True)
                opening_body = gr.Textbox(label="Opening Email", lines=12, interactive=True)
                opening_runtime = gr.Textbox(label="Runtime Status", interactive=False)

            # TAB 5: CUSTOMER END / FINDINGS
            with gr.Tab("👤 Customer End / Findings"):
                customer_issue_summary = gr.Textbox(
                    label="Issue Summary",
                    value="Reported service issue",
                )
                customer_findings = gr.CheckboxGroup(
                    choices=FINDING_OPTIONS,
                    label="Select Findings",
                )
                customer_action = gr.Dropdown(
                    choices=CUSTOMER_ACTIONS,
                    value="Verify last-mile media",
                    label="Requested Customer Action",
                )
                customer_priority = gr.Dropdown(
                    choices=["Normal", "Follow-up", "Urgent"],
                    value="Normal",
                    label="Priority",
                )

                customer_end_button = gr.Button("Generate Customer-End Response", variant="primary")
                customer_end_subject = gr.Textbox(label="Subject", interactive=True)
                customer_end_body = gr.Textbox(label="Generated Email", lines=12, interactive=True)
                customer_policy = gr.Textbox(label="Internal Policy / Warning", lines=2, interactive=False)

            # TAB 6: VENDOR / INTERNAL ESCALATION (PLACED NEXT TO CUSTOMER END)
            with gr.Tab("📨 Vendor / Internal Escalation"):
                gr.Markdown("### 📤 Vendor & Internal Team Escalation Console")

                with gr.Row():
                    escalation_format_radio = gr.Radio(
                        choices=ESCALATION_FORMAT_OPTIONS,
                        value="Option A (Standard Matrix Format)",
                        label="Email & Subject Format Mode",
                    )

                with gr.Row():
                    escalation_target = gr.Dropdown(
                        choices=ESCALATION_TARGETS,
                        value="Netsat",
                        label="Escalate To (Vendor)",
                    )
                    custom_escalation_target = gr.Textbox(
                        label="Custom Vendor Name",
                        placeholder="Used only when 'Other' is selected",
                    )
                    rcbs_group_dd = gr.Dropdown(
                        choices=["None", "RCBS North", "RCBS Central", "RCBS South"],
                        value="RCBS Central",
                        label="CC (RCBS Region Group)",
                    )

                # Editable To & CC inputs
                with gr.Row():
                    escalation_to_tb = gr.Textbox(
                        label="To: Recipients (Auto-filled L1 + L2 | Add/Edit Manual Addresses Here)",
                        value=get_vendor_l1_l2_emails("Netsat"),
                        placeholder="support.cmpak@netsat.net.pk; engineer@vendor.com",
                        lines=2,
                        interactive=True,
                    )
                    escalation_cc_tb = gr.Textbox(
                        label="Cc: Recipients (Auto-filled RCBS Group | Add/Edit Manual CC Here)",
                        value=RCBS_GROUPS.get("RCBS Central", ""),
                        placeholder="group.rcbs.central@gmail.com; supervisor@company.com",
                        lines=2,
                        interactive=True,
                    )

                # Option C Specific Custom Subject input
                with gr.Row():
                    custom_c_subject = gr.Textbox(
                        label="Custom Subject (Used primarily for Option C)",
                        placeholder="e.g. Lahore ZONG PRI DOWN vlan 345",
                        lines=1,
                    )

                with gr.Row():
                    escalation_level = gr.Dropdown(
                        choices=ESCALATION_LEVELS,
                        value="Initial engagement",
                        label="Escalation Level",
                    )
                    escalation_priority = gr.Dropdown(
                        choices=["Normal", "Follow-up", "Urgent", "Critical"],
                        value="Urgent",
                        label="Priority Tag",
                    )

                escalation_findings = gr.CheckboxGroup(
                    choices=FINDING_OPTIONS,
                    label="Current Findings (Optional)",
                )

                escalation_action = gr.Textbox(
                    label="Required Action (Optional)",
                    placeholder="e.g. Kindly dispatch a team to the site.",
                    lines=2,
                )

                escalation_button = gr.Button("🚀 Generate Escalation Email & Outlook File", variant="primary")

                escalation_subject = gr.Textbox(label="Generated Subject", interactive=True)
                escalation_body = gr.Textbox(label="Generated Email Body", lines=12, interactive=True)

                outlook_file_output = gr.File(label="📥 Download Ready Outlook File (.eml)")
                matrix_table_output = gr.HTML(label="Vendor Escalation Matrix Live Preview")

                # Dynamic Recipient Pre-Filling Events
                escalation_target.change(
                    fn=lambda v: get_vendor_l1_l2_emails(v),
                    inputs=escalation_target,
                    outputs=escalation_to_tb,
                )

                rcbs_group_dd.change(
                    fn=lambda g: RCBS_GROUPS.get(g, ""),
                    inputs=rcbs_group_dd,
                    outputs=escalation_cc_tb,
                )

            # TAB 7: CLOSURE / RFO
            with gr.Tab("✅ Closure / RFO"):
                gr.Markdown("### Close an Active Complaint")

                with gr.Row():
                    closure_complaint_selector = gr.Dropdown(
                        choices=[],
                        label="Select OPEN Complaint to Close",
                    )
                    closure_refresh_complaints = gr.Button("Refresh Open Complaint List")

                closure_selected_details = gr.Textbox(
                    label="Selected Complaint Details",
                    lines=3,
                    interactive=False,
                )

                with gr.Row():
                    issue_found_at = gr.Dropdown(
                        choices=ISSUE_FOUND_AT,
                        value="Customer",
                        label="Issue Found At",
                    )
                    custom_issue_found_at = gr.Textbox(
                        label="Custom Issue Found At",
                        placeholder="Used only for Other",
                    )

                root_cause = gr.Dropdown(
                    choices=RCA_OPTIONS,
                    value="No Issue Observed",
                    label="Root Cause",
                )
                custom_root_cause = gr.Textbox(
                    label="Custom Root Cause",
                    placeholder="Used only for Other / Manual",
                )

                corrective_action = gr.Dropdown(
                    choices=[
                        "Auto from Root Cause", "No activity performed at our end",
                        "Fiber restored", "Configuration rectified", "Link rerouted",
                        "Power restored", "Wireless parameters optimized",
                        "Optical parameters normalized", "Faulty equipment replaced",
                        "Port configuration corrected", "Issue resolved by upstream provider",
                        "Customer end issue rectified", "Link stabilized after troubleshooting",
                        "Other / Manual",
                    ],
                    value="Auto from Root Cause",
                    label="Corrective Action",
                )
                custom_corrective_action = gr.Textbox(
                    label="Custom Corrective Action",
                    placeholder="Used only for Other / Manual",
                )

                final_status = gr.Dropdown(
                    choices=[
                        "Link is up and stable.", "Service is up and working normally.",
                        "Wireless segment is up and normal.", "Link parameters are normal.",
                        "Service has been restored successfully.", "No abnormality is currently observed at our end.",
                    ],
                    value="Service is up and working normally.",
                    label="Final Status",
                )
                closure_priority = gr.Dropdown(
                    choices=["Normal", "Urgent"],
                    value="Normal",
                    label="Priority",
                )

                gr.Markdown("### 🕐 Closing / Restoration Clock")
                closure_time_mode = gr.Radio(
                    choices=[AUTO_TIME_MODE, MANUAL_TIME_MODE],
                    value=AUTO_TIME_MODE,
                    label="Closing Time Mode",
                )

                with gr.Row():
                    closure_date = gr.DateTime(
                        value=None,
                        include_time=False,
                        type="string",
                        label="Restoration Date",
                        interactive=True,
                    )
                    closure_hour = gr.Dropdown(
                        choices=HOUR_CHOICES,
                        value=None,
                        label="Hour",
                    )
                    closure_minute = gr.Dropdown(
                        choices=MINUTE_CHOICES,
                        value=None,
                        label="Minute",
                    )
                    closure_ampm = gr.Radio(
                        choices=AMPM_CHOICES,
                        value=None,
                        label="AM / PM",
                    )

                closure_set_current = gr.Button("🕐 Set Manual Picker to Current Pakistan Time")
                closure_clock_preview = gr.HTML(
                    value=clock_card_html(),
                    label="Closing Clock Preview",
                )

                closure_button = gr.Button("Generate Closure Email", variant="primary")
                closure_subject = gr.Textbox(label="Subject", interactive=True)
                closure_body = gr.Textbox(label="Generated Closure Email", lines=14, interactive=True)
                closure_runtime = gr.Textbox(label="Runtime Status", interactive=False)

            # TAB 8: TROUBLESHOOTING / STATS
            with gr.Tab("🧪 Stats / Troubleshooting"):
                stats_scenario = gr.Dropdown(
                    choices=STATS_SCENARIOS,
                    value="Packet Loss / High Latency",
                    label="Reported Scenario",
                )
                stats_audience = gr.Dropdown(
                    choices=["Customer", "Team / Vendor"],
                    value="Customer",
                    label="Audience",
                )
                stats_context = gr.Textbox(
                    label="Optional Additional Context",
                    placeholder="e.g. Please perform testing from the same affected IP pool.",
                    lines=2,
                )
                stats_button = gr.Button("Generate Required-Stats Email", variant="primary")
                stats_subject = gr.Textbox(label="Subject", interactive=True)
                stats_body = gr.Textbox(label="Generated Email", lines=14, interactive=True)
                stats_policy = gr.Textbox(label="Internal Policy / Warning", lines=2, interactive=False)

            # TAB 9: PROGRESS UPDATE
            with gr.Tab("🔄 Progress Update"):
                progress_status = gr.Dropdown(
                    choices=PROGRESS_OPTIONS,
                    value="Concerned team engaged",
                    label="Current Progress",
                )
                with gr.Row():
                    progress_ettr = gr.Dropdown(
                        choices=[
                            "Awaited", "30 Minutes", "1 Hour", "2 Hours",
                            "4 Hours", "Not Available", "Not Applicable", "Custom",
                        ],
                        value="Awaited",
                        label="ETTR",
                    )
                    progress_custom_ettr = gr.Textbox(label="Custom ETTR", placeholder="Optional")
                    progress_priority = gr.Dropdown(
                        choices=["Normal", "Follow-up", "Urgent", "Critical"],
                        value="Normal",
                        label="Priority",
                    )
                    progress_audience = gr.Dropdown(
                        choices=["Customer", "Team / Vendor"],
                        value="Customer",
                        label="Audience",
                    )

                progress_note = gr.Textbox(label="Additional Ground Update", placeholder="Optional", lines=2)
                progress_button = gr.Button("Generate Progress Update", variant="primary")
                progress_subject = gr.Textbox(label="Subject", interactive=True)
                progress_body = gr.Textbox(label="Generated Email", lines=12, interactive=True)

            # TAB 10: ACTIVE COMPLAINTS DASHBOARD
            with gr.Tab("📋 Active Complaints / Response Age"):
                with gr.Row():
                    dashboard_view = gr.Dropdown(
                        choices=[
                            "Queued + Open", "Open Only", "Queued Only",
                            "Closed Only", "All Runtime Complaints",
                        ],
                        value="Queued + Open",
                        label="View",
                    )
                    dashboard_complaint_selector = gr.Dropdown(
                        choices=[],
                        label="Select Complaint from Runtime",
                    )

                with gr.Row():
                    dashboard_refresh = gr.Button("Refresh Dashboard", variant="primary")
                    dashboard_selector_refresh = gr.Button("Refresh Complaint Selector")
                    dashboard_load_button = gr.Button("Load Selected Complaint")

                dashboard = gr.Dataframe(
                    headers=[
                        "Service", "Type", "Issue", "Added", "Opened",
                        "Age", "Closed", "Total Time", "Ticket", "Last Stage", "Status"
                    ],
                    datatype=["str"] * 11,
                    value=[],
                    interactive=False,
                    label="Runtime Complaints",
                )

                check_button = gr.Button("Check Current Label")
                incident_status = gr.Textbox(
                    label="Current Label Runtime Details",
                    lines=8,
                    interactive=False,
                )
                reset_button = gr.Button("Clear Current Label from Runtime")

        # ====================================================
        # GLOBAL EVENT BINDINGS
        # ====================================================
        add_complaints_event = add_complaints_button.click(
            fn=register_complaints,
            inputs=[bulk_complaints, complaint_default_issue, complaint_bulk_ticket],
            outputs=[complaint_selector, complaint_manager_status, complaint_manager_dashboard],
        )

        add_complaints_event.then(
            fn=refresh_opening_complaint_selector,
            inputs=[],
            outputs=opening_complaint_selector,
        ).then(
            fn=refresh_closure_complaint_selector,
            inputs=[],
            outputs=closure_complaint_selector,
        ).then(
            fn=refresh_complaint_selector,
            inputs=[],
            outputs=dashboard_complaint_selector,
        )

        refresh_complaints_button.click(
            fn=refresh_complaint_selector,
            inputs=[],
            outputs=complaint_selector,
        )

        load_complaint_button.click(
            fn=load_selected_complaint,
            inputs=complaint_selector,
            outputs=[label, common_ticket, complaint_manager_status],
        )

        opening_refresh_complaints.click(
            fn=refresh_opening_complaint_selector,
            inputs=[],
            outputs=opening_complaint_selector,
        )

        opening_complaint_selector.change(
            fn=load_selected_complaint,
            inputs=opening_complaint_selector,
            outputs=[label, common_ticket, opening_selected_details],
        )

        closure_refresh_complaints.click(
            fn=refresh_closure_complaint_selector,
            inputs=[],
            outputs=closure_complaint_selector,
        )

        closure_complaint_selector.change(
            fn=load_selected_complaint,
            inputs=closure_complaint_selector,
            outputs=[label, common_ticket, closure_selected_details],
        )

        remove_complaint_button.click(
            fn=remove_selected_complaint,
            inputs=complaint_selector,
            outputs=[complaint_selector, complaint_manager_status, complaint_manager_dashboard],
        )

        opening_set_current.click(
            fn=current_clock_picker_values,
            inputs=[],
            outputs=[
                opening_time_mode, opening_date, opening_hour,
                opening_minute, opening_ampm, opening_clock_preview,
            ],
        )

        for _comp in [opening_date, opening_hour, opening_minute, opening_ampm]:
            _comp.change(
                fn=update_clock_preview,
                inputs=[opening_date, opening_hour, opening_minute, opening_ampm],
                outputs=opening_clock_preview,
            )

        closure_set_current.click(
            fn=current_clock_picker_values,
            inputs=[],
            outputs=[
                closure_time_mode, closure_date, closure_hour,
                closure_minute, closure_ampm, closure_clock_preview,
            ],
        )

        for _comp in [closure_date, closure_hour, closure_minute, closure_ampm]:
            _comp.change(
                fn=update_clock_preview,
                inputs=[closure_date, closure_hour, closure_minute, closure_ampm],
                outputs=closure_clock_preview,
            )

        opening_event = opening_button.click(
            fn=generate_opening_email,
            inputs=[
                opening_complaint_selector, label, opening_issue, opening_custom_issue,
                common_ticket, opening_priority, opening_time_mode, opening_date,
                opening_hour, opening_minute, opening_ampm,
            ],
            outputs=[label, opening_subject, opening_body, opening_runtime],
        )

        opening_event.then(
            fn=refresh_opening_complaint_selector,
            inputs=[],
            outputs=opening_complaint_selector,
        ).then(
            fn=refresh_closure_complaint_selector,
            inputs=[],
            outputs=closure_complaint_selector,
        )

        customer_end_button.click(
            fn=generate_customer_end_email,
            inputs=[
                label, customer_issue_summary, customer_findings,
                customer_action, customer_priority, common_ticket,
            ],
            outputs=[customer_end_subject, customer_end_body, customer_policy],
        )

        # Escalation Event Binding with Manual To/CC inputs
        escalation_button.click(
            fn=generate_escalation_email,
            inputs=[
                label, escalation_target, escalation_level, escalation_format_radio,
                rcbs_group_dd, escalation_to_tb, escalation_cc_tb, custom_c_subject,
                common_ticket, escalation_findings, escalation_action,
                escalation_priority, custom_escalation_target,
            ],
            outputs=[
                escalation_subject, escalation_to_tb, escalation_cc_tb,
                escalation_body, outlook_file_output, matrix_table_output,
            ],
        )

        stats_button.click(
            fn=generate_stats_request,
            inputs=[label, stats_scenario, stats_audience, stats_context],
            outputs=[stats_subject, stats_body, stats_policy],
        )

        progress_button.click(
            fn=generate_progress_email,
            inputs=[
                label, progress_status, progress_ettr, progress_custom_ettr,
                progress_note, progress_priority, progress_audience,
            ],
            outputs=[progress_subject, progress_body],
        )

        closure_event = closure_button.click(
            fn=generate_closure_email,
            inputs=[
                closure_complaint_selector, label, issue_found_at, custom_issue_found_at,
                root_cause, custom_root_cause, corrective_action, custom_corrective_action,
                final_status, closure_priority, closure_time_mode, closure_date,
                closure_hour, closure_minute, closure_ampm,
            ],
            outputs=[label, closure_subject, closure_body, closure_runtime],
        )

        closure_event.then(
            fn=refresh_closure_complaint_selector,
            inputs=[],
            outputs=closure_complaint_selector,
        ).then(
            fn=refresh_opening_complaint_selector,
            inputs=[],
            outputs=opening_complaint_selector,
        ).then(
            fn=refresh_complaint_selector,
            inputs=[],
            outputs=dashboard_complaint_selector,
        )

        dashboard_refresh.click(
            fn=refresh_dashboard,
            inputs=dashboard_view,
            outputs=dashboard,
        )

        dashboard_selector_refresh.click(
            fn=refresh_complaint_selector,
            inputs=[],
            outputs=dashboard_complaint_selector,
        )

        dashboard_load_button.click(
            fn=load_selected_complaint,
            inputs=dashboard_complaint_selector,
            outputs=[label, common_ticket, incident_status],
        )

        check_button.click(
            fn=check_incident,
            inputs=label,
            outputs=incident_status,
        )

        reset_button.click(
            fn=reset_incident,
            inputs=label,
            outputs=incident_status,
        )

    return app




app = build_app()
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 7860))
    app.launch(server_name="0.0.0.0", server_port=port, share=False)
